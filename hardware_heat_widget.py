from __future__ import annotations

import csv
import ctypes
from ctypes import wintypes
import json
import math
import os
import queue
import re
import shutil
import subprocess
import sys
import threading
import time
import webbrowser
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional
import urllib.error
import urllib.request
import xml.etree.ElementTree as ET

import tkinter as tk
from tkinter import filedialog, messagebox, ttk


APP_NAME = "HeatLens"
APP_VERSION = "0.1.6"
POLL_INTERVAL_SECONDS = 1.5
GRAPH_X_WINDOW_LABELS: tuple[str, ...] = (
    "Auto",
    "1 minute",
    "3 minutes",
    "5 minutes",
    "10 minutes",
    "30 minutes",
)
GRAPH_Y_SCALE_LABELS: tuple[str, ...] = (
    "Auto",
    "Include zero",
    "Padded (10%)",
)
GRAPH_X_WINDOW_SECONDS: dict[str, Optional[float]] = {
    "Auto": None,
    "1 minute": 60.0,
    "3 minutes": 180.0,
    "5 minutes": 300.0,
    "10 minutes": 600.0,
    "30 minutes": 1800.0,
}
GRAPH_Y_SCALE_MODES: dict[str, str] = {
    "Auto": "auto",
    "Include zero": "zero",
    "Padded (10%)": "padded",
}
PREF_ALWAYS_START_AS_ADMIN = "always_start_as_admin"
# 1 W = 3.412141633 BTU/hr (ISO 80000-5 / NIST); same factor converts Wh -> BTU.
WATTS_TO_BTU_PER_HOUR = 3.412141633
DEFAULT_AMBIENT_TEMP_F = 72.0
DEFAULT_ROOM_VOLUME_FT3 = 1000.0
REFERENCE_EXHAUST_RISE_F = 10.0
LIBRE_HARDWARE_MONITOR_DOWNLOAD_URL = (
    "https://github.com/LibreHardwareMonitor/LibreHardwareMonitor/releases/latest"
)
# Optional tip link (Buy Me a Coffee, Ko-fi, etc.). Leave blank to hide the footer link.
DONATE_URL = "https://buymeacoffee.com/arogorn993hue"


HEAT_EQUIVALENTS: tuple[tuple[float, str], ...] = (
    (5.0,   "phone charger"),
    (10.0,  "fast phone charger"),
    (15.0,  "tablet charger"),
    (25.0,  "small network switch or router"),
    (40.0,  "desktop monitor"),
    (60.0,  "bright incandescent bulb"),
    (75.0,  "large incandescent bulb"),
    (100.0, "heated seat cushion on high"),
    (125.0, "cell phone charging + gaming"),
    (150.0, "gaming laptop light load"),
    (175.0, "gaming laptop heavy load"),
    (200.0, "1/8 space heater"),
    (300.0, "1/5 space heater"),
    (400.0, "1/4 space heater"),
    (500.0, "1/3 space heater"),
    (600.0, "2/5 space heater"),
    (750.0, "1/2 space heater"),
    (900.0, "3/5 space heater"),
    (1200.0, "4/5 space heater"),
    (1500.0, "full space heater"),
)


COLORS = {
    "bg": "#101318",
    "panel": "#151a20",
    "card": "#1b2129",
    "card_2": "#202832",
    "border": "#2b3541",
    "text": "#f4f7fb",
    "muted": "#93a1b2",
    "subtle": "#667386",
    "green": "#4fd08a",
    "amber": "#f2b84b",
    "coral": "#ff6f61",
    "cyan": "#58c7e8",
    "grid": "#26303a",
}


def set_dpi_awareness() -> None:
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def safe_float(value: object) -> Optional[float]:
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def parse_sensor_magnitude(value: object) -> Optional[float]:
    number = safe_float(value)
    if number is not None:
        return number

    text = str(value or "").replace(",", "").strip()
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return safe_float(match.group(0))


def is_live_temperature_reading(name: str, identifier: str) -> bool:
    text = f"{name} {identifier}".lower()
    non_live_terms = (
        "distance to tjmax",
        "sensor resolution",
        "low limit",
        "high limit",
        "critical low",
        "critical high",
        "warning temperature",
        "critical temperature",
    )
    return not any(term in text for term in non_live_terms)


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def is_windows_admin() -> bool:
    if sys.platform != "win32":
        return False
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return False


def launch_heatlens_elevated() -> bool:
    if sys.platform != "win32":
        return False
    try:
        if getattr(sys, "frozen", False):
            executable = sys.executable
            params = ""
            directory = str(Path(executable).resolve().parent)
        else:
            executable = sys.executable
            script = Path(__file__).resolve()
            params = f'"{script}"'
            directory = str(script.parent)
        result = ctypes.windll.shell32.ShellExecuteW(
            None,
            "runas",
            executable,
            params or None,
            directory,
            1,
        )
    except Exception:
        return False
    return int(result) > 32


def maybe_elevate_on_startup() -> bool:
    """Launch an elevated copy on Windows when requested. Returns True if this process should exit."""
    if sys.platform != "win32" or is_windows_admin():
        return False
    preferences = HeatLensPreferences()
    if not preferences.get_bool(PREF_ALWAYS_START_AS_ADMIN, False):
        return False
    return launch_heatlens_elevated()


def fahrenheit_to_celsius(value: float) -> float:
    return (value - 32.0) * 5.0 / 9.0


def celsius_to_fahrenheit(value: float) -> float:
    return value * 9.0 / 5.0 + 32.0


def format_watts(value: float) -> str:
    if value >= 100:
        return f"{value:,.0f} W"
    return f"{value:,.1f} W"


def format_kwh_per_hour(value: float) -> str:
    return f"{value / 1000.0:.3f} kWh/hr"


def format_kwh(value: float) -> str:
    if value >= 1.0:
        return f"{value:.3f} kWh"
    return f"{value * 1000.0:.1f} Wh"


def format_btu_per_hour(value: float) -> str:
    return f"{value:,.0f} BTU/hr"


def format_btu(value: float) -> str:
    if value >= 100:
        return f"{value:,.0f} BTU"
    return f"{value:,.1f} BTU"


def format_temp_f_c(fahrenheit: float) -> str:
    return f"{fahrenheit:.1f} F / {fahrenheit_to_celsius(fahrenheit):.1f} C"


def heat_equivalent_label(watts: float) -> str:
    if watts <= 0:
        return "Waiting for heat equivalent"
    equivalent = min(HEAT_EQUIVALENTS, key=lambda item: abs(item[0] - watts))
    equivalent_watts, label = equivalent
    return f"{label} ({format_watts(equivalent_watts)})"


def parse_ambient_temperature_f(value: str) -> Optional[float]:
    text = str(value or "").strip().lower()
    if not text:
        return None

    is_celsius = text.endswith("c")
    is_fahrenheit = text.endswith("f")
    if is_celsius or is_fahrenheit:
        text = text[:-1].strip()

    number = safe_float(text)
    if number is None:
        return None

    fahrenheit = celsius_to_fahrenheit(number) if is_celsius else number
    if not 0.0 <= fahrenheit <= 130.0:
        return None
    return fahrenheit


def air_heat_capacity_btu_per_ft3_f(ambient_f: float) -> float:
    absolute_rankine = ambient_f + 459.67
    density_lb_per_ft3 = 0.075 * (527.67 / absolute_rankine)
    return density_lb_per_ft3 * 0.240


def still_air_rise_f_per_hour(btu_per_hour: float, ambient_f: float, volume_ft3: float = DEFAULT_ROOM_VOLUME_FT3) -> float:
    capacity = air_heat_capacity_btu_per_ft3_f(ambient_f) * volume_ft3
    if capacity <= 0:
        return 0.0
    return btu_per_hour / capacity


def airflow_for_exhaust_rise_cfm(
    btu_per_hour: float,
    ambient_f: float,
    exhaust_rise_f: float = REFERENCE_EXHAUST_RISE_F,
) -> float:
    capacity_per_cfm = air_heat_capacity_btu_per_ft3_f(ambient_f) * 60.0 * exhaust_rise_f
    if capacity_per_cfm <= 0:
        return 0.0
    return btu_per_hour / capacity_per_cfm


def estimate_psu_efficiency(dc_watts: float) -> float:
    """Typical 80+ Gold PSU efficiency curve; low loads are less efficient."""
    if dc_watts <= 25.0:
        return 0.72
    if dc_watts <= 60.0:
        return 0.80
    if dc_watts <= 150.0:
        return 0.86
    if dc_watts <= 350.0:
        return 0.89
    if dc_watts <= 700.0:
        return 0.90
    return 0.88


def estimate_psu_loss_watts(dc_watts: float) -> float:
    if dc_watts <= 0.0:
        return 0.0
    efficiency = estimate_psu_efficiency(dc_watts)
    loss = dc_watts * (1.0 - efficiency) / efficiency
    return clamp(loss, 3.0, 120.0)


def read_sysfs_text(path: Path) -> str:
    try:
        return path.read_text(encoding="ascii").strip()
    except Exception:
        return ""


def format_temp_delta(delta_c: float) -> str:
    delta_f = delta_c * 9.0 / 5.0
    return f"{delta_c:+.1f} C / {delta_f:+.1f} F"


def format_temp(celsius: Optional[float]) -> str:
    if celsius is None:
        return "--"
    fahrenheit = celsius_to_fahrenheit(celsius)
    return f"{celsius:.0f} C / {fahrenheit:.0f} F"


def format_graph_time(timestamp: float, *, include_seconds: bool = False) -> str:
    pattern = "%I:%M:%S %p" if include_seconds else "%I:%M %p"
    return datetime.fromtimestamp(timestamp).strftime(pattern).lstrip("0")


def bus_type_label(value: Optional[float]) -> str:
    if value is None:
        return ""
    labels = {
        1: "SCSI",
        3: "ATA",
        7: "USB",
        8: "RAID",
        10: "SAS",
        11: "SATA",
        12: "SD",
        13: "MMC",
        16: "Storage Spaces",
        17: "NVMe",
        18: "SCM",
        19: "UFS",
    }
    return labels.get(int(value), "")


def media_type_label(value: Optional[float]) -> str:
    if value is None:
        return ""
    labels = {
        3: "HDD",
        4: "SSD",
        5: "SCM",
    }
    return labels.get(int(value), "")


@dataclass
class SensorReading:
    name: str
    value: float
    unit: str
    source: str
    identifier: str = ""
    estimated: bool = False
    selected_for_total: bool = False
    max_temperature_eligible: bool = True
    note: str = ""

    def display_value(self) -> str:
        prefix = "~" if self.estimated else ""
        if self.unit == "W":
            return f"{prefix}{format_watts(self.value)}"
        if self.unit == "C":
            fahrenheit = self.value * 9.0 / 5.0 + 32.0
            return f"{self.value:.1f} C / {fahrenheit:.1f} F"
        return f"{prefix}{self.value:.1f} {self.unit}"


@dataclass
class BackendResult:
    power: list[SensorReading] = field(default_factory=list)
    temperatures: list[SensorReading] = field(default_factory=list)
    estimates: list[SensorReading] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


@dataclass
class HardwareSnapshot:
    taken_at: float
    power: list[SensorReading]
    temperatures: list[SensorReading]
    selected_power: list[SensorReading]
    notes: list[str]

    @property
    def total_watts(self) -> float:
        return sum(max(0.0, reading.value) for reading in self.selected_power)

    @property
    def btu_per_hour(self) -> float:
        return self.total_watts * WATTS_TO_BTU_PER_HOUR

    @property
    def max_temp_c(self) -> Optional[float]:
        valid = [
            reading.value
            for reading in self.temperatures
            if reading.max_temperature_eligible and -40.0 <= reading.value <= 150.0
        ]
        if not valid:
            valid = [reading.value for reading in self.temperatures if -40.0 <= reading.value <= 150.0]
        return max(valid) if valid else None

    @property
    def status(self) -> str:
        if self.selected_power:
            direct = [reading for reading in self.selected_power if not reading.estimated]
            estimates = [reading for reading in self.selected_power if reading.estimated]
            if direct and estimates:
                return "Live sensors plus estimates"
            if direct:
                return "Live sensor data"
            return "Estimated wattage"
        return "Waiting for sensors"


@dataclass
class SessionLogEntry:
    timestamp: datetime
    elapsed_seconds: float
    total_watts: float
    direct_watts: float
    estimated_watts: float
    average_kwh_per_hour: Optional[float]
    btu_per_hour: float
    max_temp_c: Optional[float]
    ambient_f: Optional[float]
    heat_equivalent: str
    status: str
    selected_sources: str
    notes: str


class LibreHardwareMonitorBackend:
    name = "Libre/Open Hardware Monitor"

    def __init__(self) -> None:
        self.namespaces = ("root\\LibreHardwareMonitor", "root\\OpenHardwareMonitor")
        self.http_timeout_seconds = 1.5

    def read(self) -> BackendResult:
        http_result = self._read_http()
        if http_result.power or http_result.temperatures:
            return http_result

        wmi_result = self._read_wmi()
        if wmi_result.power or wmi_result.temperatures:
            return wmi_result

        return BackendResult(notes=compact_notes(http_result.notes + wmi_result.notes))

    def _read_http(self) -> BackendResult:
        failures: list[str] = []
        for url in self._http_urls():
            try:
                request = urllib.request.Request(url, headers={"Accept": "application/json"})
                with urllib.request.urlopen(request, timeout=self.http_timeout_seconds) as response:
                    status = getattr(response, "status", 200)
                    if status != 200:
                        failures.append(f"{url} returned HTTP {status}")
                        continue
                    payload = response.read()
            except urllib.error.URLError as exc:
                failures.append(f"{url} ({self._short_http_error(exc)})")
                continue
            except Exception as exc:
                failures.append(f"{url} ({exc})")
                continue

            try:
                data = json.loads(payload.decode("utf-8-sig"))
            except Exception as exc:
                failures.append(f"{url} returned invalid JSON ({exc})")
                continue

            result = self._read_http_tree(data, url)
            if result.power or result.temperatures:
                return result
            failures.append(f"{url} returned no power or temperature sensors")

        if not failures:
            return BackendResult()

        shown = "; ".join(failures[:3])
        extra = "" if len(failures) <= 3 else f"; plus {len(failures) - 3} more"
        return BackendResult(
            notes=[
                "Libre/OpenHardwareMonitor HTTP data.json is not reachable or has no live sensors "
                f"({shown}{extra})."
            ]
        )

    def _http_urls(self) -> list[str]:
        ports: list[int] = []
        configured_port = self._configured_http_port()
        if configured_port is not None:
            ports.append(configured_port)
        ports.extend([8085, 8086, 8080])

        unique_ports: list[int] = []
        for port in ports:
            if port not in unique_ports:
                unique_ports.append(port)

        urls: list[str] = []
        for port in unique_ports:
            urls.append(f"http://127.0.0.1:{port}/data.json")
            urls.append(f"http://localhost:{port}/data.json")
        return urls

    def _configured_http_port(self) -> Optional[int]:
        candidates = [Path(__file__).resolve().with_name("LibreHardwareMonitor.config")]
        if getattr(sys, "frozen", False):
            candidates.append(Path(sys.executable).resolve().with_name("LibreHardwareMonitor.config"))
        candidates.append(Path.cwd().resolve() / "LibreHardwareMonitor.config")

        helper = LibreHardwareMonitorHelper()
        install_path = helper.find_installation()
        if install_path is not None:
            candidates.append(install_path.with_name("LibreHardwareMonitor.config"))
        for config_exe in helper._find_via_config_files():
            candidates.append(config_exe.with_name("LibreHardwareMonitor.config"))

        seen: set[Path] = set()
        for config_path in candidates:
            if config_path in seen:
                continue
            seen.add(config_path)
            if not config_path.exists():
                continue
            try:
                root = ET.parse(config_path).getroot()
            except Exception:
                continue
            for item in root.findall(".//add"):
                if item.get("key") != "listenerPort":
                    continue
                port = safe_float(item.get("value"))
                if port is not None and 0 < int(port) <= 65535:
                    return int(port)
        return None

    def _read_http_tree(self, data: object, url: str) -> BackendResult:
        power: list[SensorReading] = []
        temperatures: list[SensorReading] = []

        for node, path in self._walk_http_nodes(data, []):
            sensor_type = str(node.get("Type") or "").strip().lower()
            if sensor_type not in ("power", "temperature"):
                continue

            value = parse_sensor_magnitude(node.get("RawValue") or node.get("Value"))
            if value is None:
                continue

            identifier = str(node.get("SensorId") or "").strip()
            name = self._http_sensor_name(path)
            source = "LibreHardwareMonitor HTTP"

            if sensor_type == "power" and 0.0 <= value <= 5000.0:
                power.append(
                    SensorReading(
                        name=name,
                        value=value,
                        unit="W",
                        source=source,
                        identifier=identifier,
                    )
                )
            elif (
                sensor_type == "temperature"
                and -40.0 <= value <= 150.0
                and is_live_temperature_reading(name, identifier)
            ):
                temperatures.append(
                    SensorReading(
                        name=name,
                        value=value,
                        unit="C",
                        source=source,
                        identifier=identifier,
                    )
                )

        return BackendResult(
            power=power,
            temperatures=temperatures,
            notes=[f"LibreHardwareMonitor HTTP data.json is available at {url}."],
        )

    def _walk_http_nodes(self, node: object, path: list[str]):
        if not isinstance(node, dict):
            return

        text = str(node.get("Text") or "").strip()
        next_path = path + [text] if text else path
        yield node, next_path

        children = node.get("Children") or []
        if not isinstance(children, list):
            return
        for child in children:
            yield from self._walk_http_nodes(child, next_path)

    def _http_sensor_name(self, path: list[str]) -> str:
        parts = [part.strip() for part in path if part and part.strip()]
        if parts and parts[0].lower() == "sensor":
            parts = parts[1:]
        if len(parts) > 1:
            parts = parts[1:]

        category_names = {
            "clocks",
            "controls",
            "data",
            "factors",
            "fans",
            "levels",
            "loads",
            "powers",
            "small data",
            "temperatures",
            "throughput",
            "voltages",
        }
        cleaned = [part for part in parts if part.lower() not in category_names]
        if cleaned:
            return " ".join(cleaned)
        if parts:
            return parts[-1]
        return "Sensor"

    def _short_http_error(self, exc: urllib.error.URLError) -> str:
        if isinstance(exc, urllib.error.HTTPError):
            return f"HTTP {exc.code}"
        reason = getattr(exc, "reason", None)
        if reason:
            return str(reason)
        return str(exc)

    def _read_wmi(self) -> BackendResult:
        try:
            import pythoncom
            import win32com.client
        except Exception:
            return BackendResult(
                notes=[
                    "pywin32 is not installed, so LibreHardwareMonitor/OpenHardwareMonitor WMI sensors are unavailable."
                ]
            )

        pythoncom.CoInitialize()
        try:
            namespace_seen = False
            for namespace in self.namespaces:
                result = self._read_namespace(win32com.client, namespace)
                namespace_seen = namespace_seen or bool(result.notes)
                if result.power or result.temperatures:
                    return result
            return BackendResult(
                notes=[
                    self._status_note(win32com.client, namespace_seen)
                ]
            )
        finally:
            pythoncom.CoUninitialize()

    def _read_namespace(self, win32com_client: object, namespace: str) -> BackendResult:
        try:
            locator = win32com_client.Dispatch("WbemScripting.SWbemLocator")
            service = locator.ConnectServer(".", namespace)
            hardware_names = self._hardware_names(service)
            sensors = service.ExecQuery(
                "SELECT * FROM Sensor WHERE SensorType='Power' OR SensorType='Temperature'"
            )
        except Exception:
            return BackendResult()

        power: list[SensorReading] = []
        temperatures: list[SensorReading] = []
        source = "LibreHardwareMonitor" if "Libre" in namespace else "OpenHardwareMonitor"

        for sensor in sensors:
            sensor_type = str(getattr(sensor, "SensorType", "")).strip()
            value = safe_float(getattr(sensor, "Value", None))
            if value is None:
                continue

            identifier = str(getattr(sensor, "Identifier", "") or "")
            name = self._sensor_name(sensor, hardware_names)

            if sensor_type.lower() == "power" and 0.0 <= value <= 5000.0:
                power.append(
                    SensorReading(
                        name=name,
                        value=value,
                        unit="W",
                        source=source,
                        identifier=identifier,
                    )
                )
            elif (
                sensor_type.lower() == "temperature"
                and -40.0 <= value <= 150.0
                and is_live_temperature_reading(name, identifier)
            ):
                temperatures.append(
                    SensorReading(
                        name=name,
                        value=value,
                        unit="C",
                        source=source,
                        identifier=identifier,
                    )
                )

        return BackendResult(power=power, temperatures=temperatures, notes=[f"{namespace} WMI namespace is available."])

    def _status_note(self, win32com_client: object, namespace_seen: bool) -> str:
        running = self._running_monitor_processes(win32com_client)
        if not running and not namespace_seen:
            return (
                "LibreHardwareMonitor/OpenHardwareMonitor is not running, and no WMI namespace is registered. "
                "Start LibreHardwareMonitor as administrator to expose motherboard, RAM, CPU, and storage sensors."
            )
        if running and not namespace_seen:
            return (
                f"{', '.join(running)} is running, but its WMI namespace is not available. "
                "Run it as administrator and enable WMI/sensor sharing if the option is present."
            )
        if namespace_seen and not running:
            return (
                "A Hardware Monitor WMI namespace exists, but LibreHardwareMonitor/OpenHardwareMonitor is not running. "
                "Start the monitor app to refresh sensor rows."
            )
        return (
            "LibreHardwareMonitor/OpenHardwareMonitor WMI is reachable, but no power or temperature sensors were returned."
        )

    def _running_monitor_processes(self, win32com_client: object) -> list[str]:
        try:
            locator = win32com_client.Dispatch("WbemScripting.SWbemLocator")
            service = locator.ConnectServer(".", "root\\cimv2")
            processes = service.ExecQuery(
                "SELECT Name FROM Win32_Process WHERE "
                "Name='LibreHardwareMonitor.exe' OR Name='OpenHardwareMonitor.exe'"
            )
        except Exception:
            return []

        names: list[str] = []
        for process in processes:
            name = str(getattr(process, "Name", "") or "").strip()
            if name and name not in names:
                names.append(name)
        return names

    def _hardware_names(self, service: object) -> dict[str, str]:
        hardware_names: dict[str, str] = {}
        try:
            hardware_items = service.ExecQuery("SELECT * FROM Hardware")
        except Exception:
            return hardware_names

        for hardware in hardware_items:
            identifier = str(getattr(hardware, "Identifier", "") or "").strip()
            name = str(getattr(hardware, "Name", "") or "").strip()
            hardware_type = str(getattr(hardware, "HardwareType", "") or "").strip()
            if not identifier:
                continue
            if hardware_type and hardware_type.lower() not in name.lower():
                name = f"{name} {hardware_type}".strip()
            hardware_names[identifier.lower()] = name or identifier
        return hardware_names

    def _sensor_name(self, sensor: object, hardware_names: dict[str, str]) -> str:
        name = str(getattr(sensor, "Name", "") or "Sensor").strip()
        parent = str(getattr(sensor, "Parent", "") or "").strip()
        identifier = str(getattr(sensor, "Identifier", "") or "").strip()

        hardware = hardware_names.get(parent.lower())
        if not hardware and identifier:
            lowered = identifier.lower()
            matches = [
                (hardware_id, hardware_name)
                for hardware_id, hardware_name in hardware_names.items()
                if lowered.startswith(hardware_id.rstrip("/") + "/")
            ]
            if matches:
                hardware = max(matches, key=lambda item: len(item[0]))[1]

        if hardware and hardware.lower() not in name.lower():
            return f"{hardware} {name}"
        if parent and parent.lower() not in name.lower() and len(parent) < 42:
            return f"{parent} {name}"
        return name


class HeatLensPreferences:
    def __init__(self) -> None:
        self.path = Path.home() / ".heatlens" / "preferences.json"
        self.data = self._load()

    def _load(self) -> dict[str, object]:
        try:
            payload = json.loads(self.path.read_text(encoding="utf-8"))
        except Exception:
            return {}
        return payload if isinstance(payload, dict) else {}

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.path.write_text(json.dumps(self.data, indent=2), encoding="utf-8")

    def get_bool(self, key: str, default: bool = False) -> bool:
        value = self.data.get(key, default)
        return bool(value)

    def set_bool(self, key: str, value: bool) -> None:
        self.data[key] = value
        self.save()


class LibreHardwareMonitorHelper:
    exe_names = ("LibreHardwareMonitor.exe", "OpenHardwareMonitor.exe")

    def __init__(self) -> None:
        self.preferences = HeatLensPreferences()
        self._http_timeout_seconds = 1.0

    def is_windows(self) -> bool:
        return sys.platform == "win32"

    def sensor_feed_available(self) -> bool:
        for port in self._candidate_ports():
            if self._http_available(port):
                return True
        return False

    def find_installation(self) -> Optional[Path]:
        if not self.is_windows():
            return None

        seen: set[Path] = set()
        for candidate in self._installation_candidates():
            try:
                resolved = candidate.resolve()
            except Exception:
                continue
            if resolved in seen or not resolved.is_file():
                continue
            seen.add(resolved)
            if resolved.name in self.exe_names:
                return resolved
        return None

    def _installation_candidates(self) -> list[Path]:
        candidates: list[Path] = []

        running = self._find_running_executable()
        if running is not None:
            candidates.append(running)

        candidates.extend(self._find_via_config_files())
        candidates.extend(self._find_via_registry())

        app_dir = Path(__file__).resolve().parent
        candidates.extend(
            [
                app_dir / "LibreHardwareMonitor.exe",
                app_dir / "OpenHardwareMonitor.exe",
            ]
        )
        if getattr(sys, "frozen", False):
            install_dir = Path(sys.executable).resolve().parent
            candidates.extend(
                [
                    install_dir / "LibreHardwareMonitor.exe",
                    install_dir / "OpenHardwareMonitor.exe",
                ]
            )

        for env_name in ("ProgramFiles", "ProgramFiles(x86)", "LOCALAPPDATA", "USERPROFILE"):
            base = os.environ.get(env_name, "")
            if not base:
                continue
            root = Path(base)
            candidates.extend(
                [
                    root / "LibreHardwareMonitor" / "LibreHardwareMonitor.exe",
                    root / "Libre Hardware Monitor" / "LibreHardwareMonitor.exe",
                    root / "Programs" / "LibreHardwareMonitor" / "LibreHardwareMonitor.exe",
                    root / "OpenHardwareMonitor" / "OpenHardwareMonitor.exe",
                    root / "Documents" / "ai" / "LibreHardwareMonitor.exe",
                    root / "OneDrive" / "Documents" / "ai" / "LibreHardwareMonitor.exe",
                    root / "Downloads" / "LibreHardwareMonitor.exe",
                    root / "Desktop" / "LibreHardwareMonitor.exe",
                ]
            )

        for env_name in ("OneDrive", "OneDriveConsumer", "OneDriveCommercial"):
            base = os.environ.get(env_name, "")
            if not base:
                continue
            root = Path(base)
            candidates.extend(
                [
                    root / "Documents" / "ai" / "LibreHardwareMonitor.exe",
                    root / "Documents" / "LibreHardwareMonitor" / "LibreHardwareMonitor.exe",
                ]
            )

        which = shutil.which("LibreHardwareMonitor")
        if which:
            candidates.append(Path(which))
        which = shutil.which("OpenHardwareMonitor")
        if which:
            candidates.append(Path(which))

        return candidates

    def _find_running_executable(self) -> Optional[Path]:
        try:
            import psutil
        except Exception:
            return None

        for process in psutil.process_iter(["name", "exe"]):
            try:
                name = str(process.info.get("name") or "")
                if name not in self.exe_names:
                    continue
                executable = process.info.get("exe")
                if executable and Path(executable).is_file():
                    return Path(executable)
            except Exception:
                continue
        return None

    def _find_via_config_files(self) -> list[Path]:
        found: list[Path] = []
        search_roots: list[Path] = []
        for value in (
            Path(__file__).resolve().parent,
            Path.cwd(),
            Path.home() / "Documents",
            Path.home() / "OneDrive" / "Documents",
            Path.home() / "Downloads",
            Path.home() / "Desktop",
        ):
            if value not in search_roots:
                search_roots.append(value)
        for env_name in ("OneDrive", "OneDriveConsumer", "OneDriveCommercial", "USERPROFILE"):
            base = os.environ.get(env_name, "")
            if not base:
                continue
            root = Path(base)
            for extra in (root, root / "Documents"):
                if extra not in search_roots:
                    search_roots.append(extra)

        seen: set[Path] = set()
        for root in search_roots:
            if not root.is_dir():
                continue
            config_paths = [root / "LibreHardwareMonitor.config"]
            try:
                for child in root.iterdir():
                    if child.is_dir():
                        config_paths.append(child / "LibreHardwareMonitor.config")
            except Exception:
                pass

            for config_path in config_paths:
                if not config_path.is_file():
                    continue
                for exe_name in self.exe_names:
                    exe_path = config_path.parent / exe_name
                    try:
                        resolved = exe_path.resolve()
                    except Exception:
                        continue
                    if resolved in seen or not resolved.is_file():
                        continue
                    seen.add(resolved)
                    found.append(resolved)
        return found

    def _find_via_registry(self) -> list[Path]:
        if sys.platform != "win32":
            return []

        try:
            import winreg
        except Exception:
            return []

        found: list[Path] = []
        seen: set[Path] = set()
        uninstall_roots = (
            (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
            (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"),
            (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
        )

        for hive, uninstall_path in uninstall_roots:
            try:
                with winreg.OpenKey(hive, uninstall_path) as uninstall_key:
                    subkey_count = winreg.QueryInfoKey(uninstall_key)[0]
                    for index in range(subkey_count):
                        try:
                            subkey_name = winreg.EnumKey(uninstall_key, index)
                            with winreg.OpenKey(uninstall_key, subkey_name) as app_key:
                                display_name = self._registry_string(app_key, "DisplayName").lower()
                                if "libre hardware" not in display_name and "open hardware" not in display_name:
                                    continue
                                for value_name in ("InstallLocation", "DisplayIcon"):
                                    raw = self._registry_string(app_key, value_name)
                                    if not raw:
                                        continue
                                    for exe_name in self.exe_names:
                                        candidate = self._path_from_registry_value(raw, exe_name)
                                        if candidate is None:
                                            continue
                                        try:
                                            resolved = candidate.resolve()
                                        except Exception:
                                            continue
                                        if resolved in seen or not resolved.is_file():
                                            continue
                                        seen.add(resolved)
                                        found.append(resolved)
                        except OSError:
                            continue
            except OSError:
                continue
        return found

    def _registry_string(self, key: object, value_name: str) -> str:
        import winreg

        try:
            value, _value_type = winreg.QueryValueEx(key, value_name)
        except OSError:
            return ""
        return str(value or "").strip().strip('"')

    def _path_from_registry_value(self, raw: str, exe_name: str) -> Optional[Path]:
        text = raw.strip().strip('"')
        if not text:
            return None

        path = Path(text)
        if path.name.lower() == exe_name.lower() and path.is_file():
            return path
        if path.suffix.lower() == ".exe" and path.is_file():
            return path
        if path.is_dir():
            candidate = path / exe_name
            if candidate.is_file():
                return candidate
        parent_candidate = path.parent / exe_name
        if parent_candidate.is_file():
            return parent_candidate
        return None

    def is_running(self) -> bool:
        if not self.is_windows():
            return False
        try:
            import psutil
        except Exception:
            return False

        for process in psutil.process_iter(["name"]):
            try:
                name = str(process.info.get("name") or "")
            except Exception:
                continue
            if name in self.exe_names:
                return True
        return False

    def __init__(self) -> None:
        self.preferences = HeatLensPreferences()
        self._http_timeout_seconds = 1.0
        self._wait_token = 0

    def ensure_web_server_config(self, install_path: Path) -> None:
        config_path = install_path.with_name("LibreHardwareMonitor.config")
        if not config_path.is_file():
            return
        try:
            text = config_path.read_text(encoding="utf-8")
        except Exception:
            return

        updated = text
        if 'key="runWebServerMenuItem"' in updated:
            updated = re.sub(
                r'(<add key="runWebServerMenuItem" value=")([^"]*)(" />)',
                r"\1true\3",
                updated,
                count=1,
            )
        else:
            updated = updated.replace(
                "</appSettings>",
                '    <add key="runWebServerMenuItem" value="true" />\n  </appSettings>',
                1,
            )

        if updated != text:
            try:
                config_path.write_text(updated, encoding="utf-8")
            except Exception:
                pass

    def launch(self, install_path: Path) -> bool:
        self.ensure_web_server_config(install_path)
        try:
            subprocess.Popen(
                [str(install_path)],
                cwd=str(install_path.parent),
            )
        except Exception:
            return False
        return True

    def launch_elevated(self, install_path: Path) -> bool:
        if sys.platform != "win32":
            return False
        self.ensure_web_server_config(install_path)
        try:
            result = ctypes.windll.shell32.ShellExecuteW(
                None,
                "runas",
                str(install_path),
                None,
                str(install_path.parent),
                1,
            )
        except Exception:
            return False
        return int(result) > 32

    def begin_sensor_feed_wait(
        self,
        parent: tk.Misc,
        install_path: Path,
        *,
        on_status: Callable[[str], None],
        on_connected: Callable[[], None],
        on_failed: Callable[[], None],
        launch_if_needed: bool = True,
        attempt: int = 0,
        max_attempts: int = 30,
        wait_token: int = 0,
    ) -> int:
        self._wait_token += 1
        token = self._wait_token
        parent.after(
            0,
            lambda: self._poll_sensor_feed_wait(
                parent,
                install_path,
                on_status=on_status,
                on_connected=on_connected,
                on_failed=on_failed,
                launch_if_needed=launch_if_needed,
                attempt=attempt,
                max_attempts=max_attempts,
                wait_token=token,
            ),
        )
        return token

    def _poll_sensor_feed_wait(
        self,
        parent: tk.Misc,
        install_path: Path,
        *,
        on_status: Callable[[str], None],
        on_connected: Callable[[], None],
        on_failed: Callable[[], None],
        launch_if_needed: bool,
        attempt: int,
        max_attempts: int,
        wait_token: int,
    ) -> None:
        if wait_token != self._wait_token:
            return

        if self.sensor_feed_available():
            on_connected()
            return

        if attempt == 0 and launch_if_needed and not self.is_running():
            on_status("Starting LibreHardwareMonitor...")
            if not self.launch(install_path):
                on_failed()
                return
        elif attempt == 0 and self.is_running():
            on_status("LibreHardwareMonitor is running — waiting for sensor feed...")

        if attempt >= max_attempts:
            on_failed()
            return

        if not self.is_running() and attempt > 2:
            on_status("LibreHardwareMonitor closed before sensors connected.")
            on_failed()
            return

        if self.is_running():
            on_status(
                "Waiting for Libre sensors... enable Options -> Remote Web Server -> Run "
                f"({attempt + 1}/{max_attempts})"
            )
        else:
            on_status(f"Waiting for LibreHardwareMonitor to start ({attempt + 1}/{max_attempts})")

        parent.after(
            2000,
            lambda: self._poll_sensor_feed_wait(
                parent,
                install_path,
                on_status=on_status,
                on_connected=on_connected,
                on_failed=on_failed,
                launch_if_needed=False,
                attempt=attempt + 1,
                max_attempts=max_attempts,
                wait_token=wait_token,
            ),
        )

    def open_download_page(self) -> None:
        webbrowser.open(LIBRE_HARDWARE_MONITOR_DOWNLOAD_URL)

    def open_donate_page(self) -> None:
        if DONATE_URL.strip():
            webbrowser.open(DONATE_URL.strip())

    def maybe_prompt_on_startup(
        self,
        parent: tk.Misc,
        *,
        on_status: Callable[[str], None],
        on_connected: Callable[[], None],
        on_failed: Callable[[], None],
    ) -> None:
        if not self.is_windows():
            return
        if self.preferences.get_bool("suppress_libre_setup_prompt"):
            return
        if self.sensor_feed_available():
            return

        install_path = self.find_installation()
        if install_path is not None:
            if self.is_running():
                self.begin_sensor_feed_wait(
                    parent,
                    install_path,
                    on_status=on_status,
                    on_connected=on_connected,
                    on_failed=on_failed,
                    launch_if_needed=False,
                )
                return
            if messagebox.askyesno(
                "Better sensor coverage",
                "LibreHardwareMonitor is installed but not running.\n\n"
                "HeatLens can use it for motherboard, CPU package, RAM, and storage sensors.\n\n"
                "Start LibreHardwareMonitor now?",
                parent=parent,
            ):
                self.begin_sensor_feed_wait(
                    parent,
                    install_path,
                    on_status=on_status,
                    on_connected=on_connected,
                    on_failed=on_failed,
                    launch_if_needed=True,
                )
            return

        response = messagebox.askyesnocancel(
            "Better sensor coverage",
            "LibreHardwareMonitor is not installed on this PC.\n\n"
            "HeatLens works without it, but Libre gives the best wattage and temperature coverage "
            "for CPU, GPU, motherboard, RAM, and NVMe.\n\n"
            "Open the download page?",
            parent=parent,
        )
        if response is None:
            self.preferences.set_bool("suppress_libre_setup_prompt", True)
        elif response:
            self.open_download_page()

    def prompt_manual_action(
        self,
        parent: tk.Misc,
        *,
        on_status: Callable[[str], None],
        on_connected: Callable[[], None],
        on_failed: Callable[[], None],
    ) -> None:
        if not self.is_windows():
            messagebox.showinfo(
                "LibreHardwareMonitor",
                "LibreHardwareMonitor is a Windows app. On Linux, use RAPL/hwmon sensors or run "
                "LibreHardwareMonitor under Wine if you need it.",
                parent=parent,
            )
            return

        if self.sensor_feed_available():
            messagebox.showinfo(
                "LibreHardwareMonitor",
                "HeatLens is already receiving LibreHardwareMonitor sensor data.",
                parent=parent,
            )
            return

        install_path = self.find_installation()
        if install_path is None:
            if messagebox.askyesno(
                "LibreHardwareMonitor",
                "LibreHardwareMonitor was not found on this PC.\n\nOpen the download page?",
                parent=parent,
            ):
                self.open_download_page()
            return

        if self.is_running():
            self.begin_sensor_feed_wait(
                parent,
                install_path,
                on_status=on_status,
                on_connected=on_connected,
                on_failed=on_failed,
                launch_if_needed=False,
            )
            return

        if messagebox.askyesno(
            "LibreHardwareMonitor",
            f"Found LibreHardwareMonitor at:\n{install_path}\n\nStart it now?",
            parent=parent,
        ):
            self.begin_sensor_feed_wait(
                parent,
                install_path,
                on_status=on_status,
                on_connected=on_connected,
                on_failed=on_failed,
                launch_if_needed=True,
            )

    def _prompt_enable_web_server(self, parent: tk.Misc) -> None:
        if self.sensor_feed_available():
            return
        messagebox.showinfo(
            "Enable LibreHardwareMonitor web server",
            "LibreHardwareMonitor is running, but HeatLens still cannot read its sensor feed.\n\n"
            "In the LibreHardwareMonitor window:\n"
            "1. Open Options\n"
            "2. Open Remote Web Server\n"
            "3. Click Run (checkmark should appear)\n\n"
            "Default port is 8085. After enabling Run, HeatLens should update within a few seconds.\n\n"
            "Tip: run LibreHardwareMonitor as administrator for the fullest sensor set.",
            parent=parent,
        )
        if messagebox.askyesno(
            "Try administrator launch?",
            "If the web server still will not start, try launching LibreHardwareMonitor as administrator.\n\n"
            "Do that now?",
            parent=parent,
        ):
            install_path = self.find_installation()
            if install_path is not None:
                self.launch_elevated(install_path)

    def _candidate_ports(self) -> list[int]:
        ports: list[int] = []
        configured = LibreHardwareMonitorBackend()._configured_http_port()
        if configured is not None:
            ports.append(configured)
        for port in (8085, 8086, 8080):
            if port not in ports:
                ports.append(port)
        return ports

    def _http_available(self, port: int) -> bool:
        url = f"http://127.0.0.1:{port}/data.json"
        try:
            request = urllib.request.Request(url, headers={"Accept": "application/json"})
            with urllib.request.urlopen(request, timeout=self._http_timeout_seconds) as response:
                if getattr(response, "status", 200) != 200:
                    return False
                payload = response.read(4096)
        except Exception:
            return False
        return bool(payload.strip())


class NvidiaSmiBackend:
    name = "nvidia-smi"

    def __init__(self) -> None:
        self.executable = self._find_executable()

    def read(self) -> BackendResult:
        if not self.executable:
            return BackendResult()

        command = [
            self.executable,
            "--query-gpu=index,name,power.draw,temperature.gpu",
            "--format=csv,noheader,nounits",
        ]
        try:
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=2.5,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
        except Exception as exc:
            return BackendResult(notes=[f"nvidia-smi could not be read: {exc}"])

        if completed.returncode != 0:
            return BackendResult()

        power: list[SensorReading] = []
        temperatures: list[SensorReading] = []
        for row in csv.reader(completed.stdout.splitlines()):
            if len(row) < 4:
                continue
            index = row[0].strip()
            gpu_name = row[1].strip()
            watts = safe_float(row[2])
            temp_c = safe_float(row[3])
            label = f"GPU {index} {gpu_name}".strip()

            if watts is not None and 0.0 <= watts <= 1200.0:
                power.append(
                    SensorReading(
                        name=f"{label} power draw",
                        value=watts,
                        unit="W",
                        source="nvidia-smi",
                        identifier=f"nvidia-smi:gpu:{index}:power",
                    )
                )
            if temp_c is not None and -40.0 <= temp_c <= 150.0:
                temperatures.append(
                    SensorReading(
                        name=f"{label} core temperature",
                        value=temp_c,
                        unit="C",
                        source="nvidia-smi",
                        identifier=f"nvidia-smi:gpu:{index}:temperature",
                    )
                )

        return BackendResult(power=power, temperatures=temperatures)

    def _find_executable(self) -> str:
        found = shutil.which("nvidia-smi")
        if found:
            return found
        if sys.platform == "win32":
            candidate = r"C:\Program Files\NVIDIA Corporation\NVSMI\nvidia-smi.exe"
            if shutil.which(candidate) or __import__("os").path.exists(candidate):
                return candidate
        return ""


class LinuxRaplBackend:
    name = "Linux RAPL"

    def __init__(self) -> None:
        self._last_samples: dict[str, tuple[int, float]] = {}

    def read(self) -> BackendResult:
        if sys.platform != "linux":
            return BackendResult()

        powercap_root = Path("/sys/class/powercap")
        if not powercap_root.is_dir():
            return BackendResult(notes=["Linux RAPL powercap sysfs is not available on this kernel."])

        power: list[SensorReading] = []
        now = time.monotonic()
        for domain_path in sorted(powercap_root.iterdir()):
            energy_file = domain_path / "energy_uj"
            if not energy_file.exists():
                continue

            domain_name = read_sysfs_text(domain_path / "name") or domain_path.name
            if not self._is_package_domain(domain_name):
                continue

            try:
                energy_uj = int(read_sysfs_text(energy_file))
            except ValueError:
                continue

            key = str(domain_path)
            previous = self._last_samples.get(key)
            self._last_samples[key] = (energy_uj, now)
            if previous is None:
                continue

            last_energy, last_time = previous
            delta_uj = energy_uj - last_energy
            if delta_uj < 0:
                try:
                    max_uj = int(read_sysfs_text(domain_path / "max_energy_range_uj"))
                except ValueError:
                    max_uj = 0
                if max_uj > 0:
                    delta_uj = (max_uj - last_energy) + energy_uj
                else:
                    continue

            delta_t = now - last_time
            if delta_t <= 0.0 or delta_uj <= 0:
                continue

            watts = (delta_uj / 1_000_000.0) / delta_t
            if not 0.0 <= watts <= 1200.0:
                continue

            power.append(
                SensorReading(
                    name=f"CPU {domain_name} power",
                    value=watts,
                    unit="W",
                    source="Linux RAPL",
                    identifier=f"linux-rapl:{domain_path.name}:power",
                )
            )

        if power:
            return BackendResult(power=power, notes=["CPU package power is from Linux RAPL energy counters."])
        return BackendResult(notes=["Linux RAPL is present, but package power needs one sample interval to settle."])

    def _is_package_domain(self, name: str) -> bool:
        lowered = name.lower()
        if "package" in lowered or lowered.startswith("psys"):
            return True
        return lowered in {"cpu", "processor"}


class LinuxHwmonBackend:
    name = "Linux hwmon"

    def read(self) -> BackendResult:
        if sys.platform != "linux":
            return BackendResult()

        power: list[SensorReading] = []
        temperatures: list[SensorReading] = []
        hwmon_root = Path("/sys/class/hwmon")
        if not hwmon_root.is_dir():
            return BackendResult()

        for hwmon_path in sorted(hwmon_root.iterdir()):
            chip_name = read_sysfs_text(hwmon_path / "name") or hwmon_path.name
            for input_path in sorted(hwmon_path.glob("*_input")):
                stem = input_path.stem
                if stem.startswith("power"):
                    watts = self._read_hwmon_power_watts(input_path)
                    if watts is None:
                        continue
                    label = read_sysfs_text(hwmon_path / stem.replace("_input", "_label")) or stem
                    power.append(
                        SensorReading(
                            name=f"{chip_name} {label}".strip(),
                            value=watts,
                            unit="W",
                            source="Linux hwmon",
                            identifier=f"linux-hwmon:{hwmon_path.name}:{stem}",
                        )
                    )
                elif stem.startswith("temp"):
                    temp_c = self._read_hwmon_temp_c(input_path)
                    if temp_c is None:
                        continue
                    label = read_sysfs_text(hwmon_path / stem.replace("_input", "_label")) or stem
                    name = f"{chip_name} {label}".strip()
                    if not is_live_temperature_reading(name, stem):
                        continue
                    temperatures.append(
                        SensorReading(
                            name=name,
                            value=temp_c,
                            unit="C",
                            source="Linux hwmon",
                            identifier=f"linux-hwmon:{hwmon_path.name}:{stem}",
                        )
                    )

        return BackendResult(power=power, temperatures=temperatures)

    def _read_hwmon_power_watts(self, path: Path) -> Optional[float]:
        try:
            raw = int(read_sysfs_text(path))
        except ValueError:
            return None
        watts = raw / 1_000_000.0
        if not 0.0 <= watts <= 1200.0:
            return None
        return watts

    def _read_hwmon_temp_c(self, path: Path) -> Optional[float]:
        try:
            raw = int(read_sysfs_text(path))
        except ValueError:
            return None
        if raw <= 0:
            return None
        temp_c = raw / 1000.0
        if not -40.0 <= temp_c <= 150.0:
            return None
        return temp_c


class RocmSmiBackend:
    name = "rocm-smi"

    def __init__(self) -> None:
        self.executable = shutil.which("rocm-smi") or ""

    def read(self) -> BackendResult:
        if not self.executable:
            return BackendResult()

        command = [self.executable, "--showpower", "--showtemp", "--json"]
        try:
            completed = subprocess.run(command, capture_output=True, text=True, timeout=3.0)
        except Exception as exc:
            return BackendResult(notes=[f"rocm-smi could not be read: {exc}"])

        if completed.returncode != 0 or not completed.stdout.strip():
            return BackendResult()

        try:
            payload = json.loads(completed.stdout)
        except json.JSONDecodeError:
            return BackendResult()

        power: list[SensorReading] = []
        temperatures: list[SensorReading] = []
        cards = payload.get("card") or payload.get("cards") or payload
        if isinstance(cards, dict):
            items = cards.items()
        elif isinstance(cards, list):
            items = enumerate(cards)
        else:
            return BackendResult()

        for index, card in items:
            if not isinstance(card, dict):
                continue
            label = str(card.get("Card series") or card.get("Card model") or f"GPU {index}").strip()
            watts = safe_float(card.get("Average Graphics Package Power (W)") or card.get("Current Socket Graphics Package Power (W)"))
            temp_c = safe_float(card.get("Temperature (Sensor edge) (C)") or card.get("Temperature (Sensor junction) (C)"))
            if watts is not None and 0.0 <= watts <= 1200.0:
                power.append(
                    SensorReading(
                        name=f"{label} power draw",
                        value=watts,
                        unit="W",
                        source="rocm-smi",
                        identifier=f"rocm-smi:gpu:{index}:power",
                    )
                )
            if temp_c is not None and -40.0 <= temp_c <= 150.0:
                temperatures.append(
                    SensorReading(
                        name=f"{label} core temperature",
                        value=temp_c,
                        unit="C",
                        source="rocm-smi",
                        identifier=f"rocm-smi:gpu:{index}:temperature",
                    )
                )

        return BackendResult(power=power, temperatures=temperatures)


class WindowsNativeTemperatureBackend:
    name = "Windows native sensors"

    def read(self) -> BackendResult:
        if sys.platform != "win32":
            return BackendResult()
        try:
            import pythoncom
            import win32com.client
        except Exception:
            return BackendResult(
                notes=["pywin32 is not installed, so native Windows WMI temperatures are unavailable."]
            )

        pythoncom.CoInitialize()
        try:
            temperatures: list[SensorReading] = []
            notes: list[str] = []
            for label, reader in (
                ("ACPI thermal zone", self._read_acpi_zones),
                ("Windows storage", self._read_storage_temperatures),
            ):
                try:
                    result = reader(win32com.client)
                except Exception as exc:
                    notes.append(self._friendly_wmi_error(label, exc))
                    continue
                temperatures.extend(result.temperatures)
                notes.extend(result.notes)
            return BackendResult(temperatures=temperatures, notes=notes)
        finally:
            pythoncom.CoUninitialize()

    def _read_acpi_zones(self, win32com_client: object) -> BackendResult:
        try:
            locator = win32com_client.Dispatch("WbemScripting.SWbemLocator")
            service = locator.ConnectServer(".", "root\\WMI")
            zones = service.ExecQuery(
                "SELECT InstanceName, CurrentTemperature FROM MSAcpi_ThermalZoneTemperature"
            )
        except Exception:
            return BackendResult()

        temperatures: list[SensorReading] = []
        try:
            zone_items = list(zones)
        except Exception as exc:
            return BackendResult(notes=[self._friendly_wmi_error("ACPI thermal zone", exc)])

        for index, zone in enumerate(zone_items):
            raw_temp = safe_float(getattr(zone, "CurrentTemperature", None))
            if raw_temp is None:
                continue
            temp_c = raw_temp / 10.0 - 273.15
            if not (-40.0 <= temp_c <= 150.0):
                continue
            instance = str(getattr(zone, "InstanceName", "") or "").strip()
            label = self._clean_acpi_name(instance) or f"Thermal zone {index + 1}"
            temperatures.append(
                SensorReading(
                    name=f"ACPI {label}",
                    value=temp_c,
                    unit="C",
                    source="Windows ACPI",
                    identifier=f"windows-acpi:{instance or index}",
                )
            )
        return BackendResult(temperatures=temperatures)

    def _read_storage_temperatures(self, win32com_client: object) -> BackendResult:
        try:
            locator = win32com_client.Dispatch("WbemScripting.SWbemLocator")
            service = locator.ConnectServer(".", "root\\Microsoft\\Windows\\Storage")
        except Exception:
            return BackendResult()

        disk_names = self._read_physical_disk_names(service)
        try:
            counters = service.ExecQuery(
                "SELECT DeviceId, Temperature FROM MSFT_StorageReliabilityCounter"
            )
        except Exception:
            if disk_names:
                return BackendResult(
                    notes=[
                        "Windows can see physical disks, but storage temperature counters are unavailable. Running as administrator or drive firmware support may be required."
                    ]
                )
            return BackendResult()

        temperatures: list[SensorReading] = []
        try:
            counter_items = list(counters)
        except Exception as exc:
            if disk_names:
                return BackendResult(notes=[self._friendly_wmi_error("Windows storage", exc)])
            return BackendResult()

        for counter in counter_items:
            device_id = str(getattr(counter, "DeviceId", "") or "").strip()
            temp_c = safe_float(getattr(counter, "Temperature", None))
            if temp_c is None or not (1.0 <= temp_c <= 150.0):
                continue
            disk_name = disk_names.get(device_id, f"Physical disk {device_id}".strip())
            temperatures.append(
                SensorReading(
                    name=f"{disk_name} storage temperature",
                    value=temp_c,
                    unit="C",
                    source="Windows Storage",
                    identifier=f"windows-storage:physical-disk:{device_id}:temperature",
                )
            )
        return BackendResult(temperatures=temperatures)

    def _read_physical_disk_names(self, service: object) -> dict[str, str]:
        disk_names: dict[str, str] = {}
        try:
            disks = service.ExecQuery(
                "SELECT DeviceId, FriendlyName, MediaType, BusType FROM MSFT_PhysicalDisk"
            )
        except Exception:
            return disk_names

        for disk in disks:
            device_id = str(getattr(disk, "DeviceId", "") or "").strip()
            friendly_name = str(getattr(disk, "FriendlyName", "") or "").strip()
            media = media_type_label(safe_float(getattr(disk, "MediaType", None)))
            bus = bus_type_label(safe_float(getattr(disk, "BusType", None)))
            parts = [part for part in (friendly_name, media, bus) if part]
            if device_id:
                disk_names[device_id] = " ".join(parts) or f"Physical disk {device_id}"
        return disk_names

    def _clean_acpi_name(self, instance: str) -> str:
        if not instance:
            return ""
        label = instance.split("\\")[-1].split("_0")[-1]
        label = label.replace("_", " ").replace("#", " ").strip()
        return " ".join(label.split())

    def _friendly_wmi_error(self, label: str, exc: Exception) -> str:
        text = str(exc)
        if "0x80041003" in text or "-2147217405" in text or "WBEM_E_ACCESS_DENIED" in text:
            return f"{label} sensors are blocked by WMI permissions; run HeatLens as administrator to try those readings."
        return f"{label} sensors are unavailable through Windows WMI."


class WindowsCpuLoadTracker:
    def __init__(self) -> None:
        self.last_idle: Optional[int] = None
        self.last_total: Optional[int] = None

    def percent(self) -> Optional[float]:
        if sys.platform != "win32":
            return None

        idle = wintypes.FILETIME()
        kernel = wintypes.FILETIME()
        user = wintypes.FILETIME()
        try:
            ok = ctypes.windll.kernel32.GetSystemTimes(
                ctypes.byref(idle), ctypes.byref(kernel), ctypes.byref(user)
            )
        except Exception:
            return None
        if not ok:
            return None

        idle_time = self._filetime_to_int(idle)
        total_time = self._filetime_to_int(kernel) + self._filetime_to_int(user)
        if self.last_idle is None or self.last_total is None:
            self.last_idle = idle_time
            self.last_total = total_time
            return None

        idle_delta = idle_time - self.last_idle
        total_delta = total_time - self.last_total
        self.last_idle = idle_time
        self.last_total = total_time
        if total_delta <= 0:
            return None

        busy_delta = max(0, total_delta - idle_delta)
        return clamp((busy_delta / total_delta) * 100.0, 0.0, 100.0)

    def _filetime_to_int(self, filetime: wintypes.FILETIME) -> int:
        return (int(filetime.dwHighDateTime) << 32) + int(filetime.dwLowDateTime)


class PsutilEstimateBackend:
    name = "psutil"

    def __init__(self) -> None:
        self.windows_cpu_load = WindowsCpuLoadTracker()
        try:
            import psutil
        except Exception:
            self.psutil = None
        else:
            self.psutil = psutil
            try:
                self.psutil.cpu_percent(interval=None)
            except Exception:
                pass

    def read(self) -> BackendResult:
        notes: list[str] = []
        if self.psutil is None:
            notes.append(
                "Install psutil for extra OS temperature support; CPU/platform estimates are using Windows CPU load."
            )

        temperatures = self._read_temperatures() if self.psutil is not None else []
        cpu_percent = self._read_cpu_percent()
        max_temp = max((sensor.value for sensor in temperatures), default=None)
        estimates = []

        if cpu_percent is not None:
            cpu_load = clamp(cpu_percent / 100.0, 0.0, 1.0)
            temp_bonus = 0.0
            if max_temp is not None:
                temp_bonus = clamp(max_temp - 55.0, 0.0, 35.0) * 0.55

            cpu_platform_watts = 20.0 + 105.0 * (cpu_load**1.35) + temp_bonus
            system_watts = 42.0 + 135.0 * (cpu_load**1.25) + temp_bonus

            estimates.append(
                SensorReading(
                    name="Estimated CPU/platform power",
                    value=cpu_platform_watts,
                    unit="W",
                    source=self._estimate_source(),
                    identifier="estimate:cpu-platform",
                    estimated=True,
                )
            )
            estimates.append(
                SensorReading(
                    name="Estimated system power",
                    value=system_watts,
                    unit="W",
                    source=self._estimate_source(),
                    identifier="estimate:system",
                    estimated=True,
                )
            )

        return BackendResult(temperatures=temperatures, estimates=estimates, notes=notes)

    def _read_cpu_percent(self) -> Optional[float]:
        if self.psutil is None:
            return self.windows_cpu_load.percent()
        try:
            return float(self.psutil.cpu_percent(interval=None))
        except Exception:
            return None

    def _estimate_source(self) -> str:
        if self.psutil is None:
            return "Windows CPU load estimate"
        return "psutil estimate"

    def _read_temperatures(self) -> list[SensorReading]:
        temperatures: list[SensorReading] = []
        try:
            temp_groups = self.psutil.sensors_temperatures(fahrenheit=False)
        except Exception:
            return temperatures

        for chip_name, entries in temp_groups.items():
            for index, entry in enumerate(entries):
                value = safe_float(getattr(entry, "current", None))
                if value is None or not (-40.0 <= value <= 150.0):
                    continue
                label = str(getattr(entry, "label", "") or "").strip()
                name = f"{chip_name} {label}".strip() or f"Temperature {index + 1}"
                temperatures.append(
                    SensorReading(
                        name=name,
                        value=value,
                        unit="C",
                        source="psutil",
                        identifier=f"psutil:{chip_name}:{index}",
                    )
                )
        return temperatures


class SensorCollector:
    def __init__(self) -> None:
        self.backends = [
            LibreHardwareMonitorBackend(),
            LinuxRaplBackend(),
            LinuxHwmonBackend(),
            NvidiaSmiBackend(),
            RocmSmiBackend(),
            WindowsNativeTemperatureBackend(),
            PsutilEstimateBackend(),
        ]

    def sample(self) -> HardwareSnapshot:
        power: list[SensorReading] = []
        temperatures: list[SensorReading] = []
        estimates: list[SensorReading] = []
        notes: list[str] = []

        for backend in self.backends:
            try:
                result = backend.read()
            except Exception as exc:
                notes.append(f"{backend.name}: {exc}")
                continue
            power.extend(result.power)
            temperatures.extend(result.temperatures)
            estimates.extend(result.estimates)
            notes.extend(result.notes)

        direct_power = dedupe_readings(power)
        temperatures = dedupe_readings(temperatures)
        temperature_notes = annotate_pascal_gpu_hotspots(temperatures)
        if temperature_notes:
            notes = temperature_notes + notes
        selected = select_power_for_total(direct_power)
        selected_groups = {power_group_key(reading) for reading in selected}

        if not selected:
            system_estimate = find_estimate(estimates, "estimate:system")
            if system_estimate:
                system_estimate.selected_for_total = True
                selected.append(system_estimate)
                notes.append("Total wattage is estimated because direct power sensors were not found.")
        elif not any(group.startswith("cpu:") for group in selected_groups):
            cpu_estimate = find_estimate(estimates, "estimate:cpu-platform")
            if cpu_estimate:
                cpu_estimate.selected_for_total = True
                selected.append(cpu_estimate)
                notes.append("CPU/platform wattage is estimated because no CPU package power sensor was found.")
            else:
                notes.append(
                    "No CPU package power sensor is available, so total wattage may only reflect visible power sensors."
                )

        coverage_estimates = estimate_unmetered_power(selected, temperatures)
        if coverage_estimates:
            estimates.extend(coverage_estimates)
            selected.extend(coverage_estimates)
            notes.append(
                "Total wattage includes estimated motherboard/RAM/storage/PSU overhead because direct sensors do not cover the whole PC."
            )

        for reading in direct_power:
            reading.selected_for_total = any(reading is selected_item for selected_item in selected)

        all_power_for_display = list(direct_power)
        for estimate in estimates:
            if any(estimate is selected_item for selected_item in selected):
                all_power_for_display.append(estimate)

        if not selected:
            notes.append(
                "No wattage source is available yet. Install psutil or start LibreHardwareMonitor/OpenHardwareMonitor with WMI enabled."
            )

        return HardwareSnapshot(
            taken_at=time.time(),
            power=sorted(all_power_for_display, key=lambda reading: (not reading.selected_for_total, reading.name.lower())),
            temperatures=sorted(temperatures, key=lambda reading: reading.name.lower()),
            selected_power=selected,
            notes=compact_notes(notes),
        )


def dedupe_readings(readings: list[SensorReading]) -> list[SensorReading]:
    deduped: dict[tuple[str, str, str], SensorReading] = {}
    for reading in readings:
        key = (
            reading.source.lower(),
            reading.identifier.lower() or reading.name.lower(),
            reading.unit,
        )
        existing = deduped.get(key)
        if existing is None or abs(reading.value) > abs(existing.value):
            deduped[key] = reading
    return list(deduped.values())


def compact_notes(notes: list[str]) -> list[str]:
    seen: set[str] = set()
    compacted: list[str] = []
    for note in notes:
        note = " ".join(str(note).split())
        if not note or note in seen:
            continue
        seen.add(note)
        compacted.append(note)
    return compacted[:4]


def find_estimate(estimates: list[SensorReading], identifier: str) -> Optional[SensorReading]:
    for estimate in estimates:
        if estimate.identifier == identifier:
            return estimate
    return None


def estimate_unmetered_power(selected_power: list[SensorReading], temperatures: list[SensorReading]) -> list[SensorReading]:
    measured_watts = sum(max(0.0, reading.value) for reading in selected_power)
    if measured_watts <= 0.0 or has_whole_system_power_sensor(selected_power):
        return []

    estimates: list[SensorReading] = []
    platform_watts = estimate_platform_watts(temperatures)
    ram_watts = estimate_ram_watts(temperatures)
    storage_watts = estimate_storage_watts(temperatures)

    if platform_watts > 0.0:
        estimates.append(
            SensorReading(
                name="Estimated motherboard, chipset, fans, and USB power",
                value=platform_watts,
                unit="W",
                source="coverage estimate",
                identifier="estimate:unmetered-platform",
                estimated=True,
                selected_for_total=True,
            )
        )
    if ram_watts > 0.0:
        estimates.append(
            SensorReading(
                name=estimate_ram_label(temperatures),
                value=ram_watts,
                unit="W",
                source="coverage estimate",
                identifier="estimate:ram-dimms",
                estimated=True,
                selected_for_total=True,
            )
        )
    if storage_watts > 0.0:
        estimates.append(
            SensorReading(
                name="Estimated NVMe/storage power",
                value=storage_watts,
                unit="W",
                source="coverage estimate",
                identifier="estimate:storage",
                estimated=True,
                selected_for_total=True,
            )
        )

    dc_watts = measured_watts + sum(reading.value for reading in estimates)
    psu_loss_watts = estimate_psu_loss_watts(dc_watts)
    estimates.append(
        SensorReading(
            name="Estimated PSU conversion loss",
            value=psu_loss_watts,
            unit="W",
            source="coverage estimate",
            identifier="estimate:psu-loss",
            estimated=True,
            selected_for_total=True,
        )
    )
    return estimates


def has_whole_system_power_sensor(readings: list[SensorReading]) -> bool:
    for reading in readings:
        text = f"{reading.name} {reading.identifier} {reading.source}".lower()
        if reading.identifier == "estimate:system":
            return True
        if reading.estimated:
            continue
        if "wall" in text or "ups" in text or "whole system" in text or "computer total" in text:
            return True
    return False


def estimate_platform_watts(temperatures: list[SensorReading]) -> float:
    platform_temps = [
        reading.value
        for reading in temperatures
        if any(
            token in f"{reading.name} {reading.identifier}".lower()
            for token in ("lpc", "motherboard", "chipset", "vrm", "mos", "pcie x16", "system #")
        )
    ]
    if not platform_temps:
        return 24.0

    max_temp = max(platform_temps)
    watts = 26.0
    if max_temp >= 55.0:
        watts += 4.0
    if max_temp >= 65.0:
        watts += 6.0
    return watts


def ram_memory_type_label(temperatures: list[SensorReading]) -> str:
    for reading in temperatures:
        text = f"{reading.name} {reading.identifier}".lower()
        if "ddr5" in text:
            return "DDR5"
        if "ddr4" in text:
            return "DDR4"
        if "lpddr" in text:
            return "LPDDR"
    return "DDR"


def estimate_ram_watts(temperatures: list[SensorReading]) -> float:
    dimms: dict[str, float] = {}
    for reading in temperatures:
        text = f"{reading.name} {reading.identifier}".lower()
        if "/memory/dimm/" not in text and "dimm" not in text and "memory" not in text:
            continue
        match = re.search(r"/memory/dimm/([^/]+)", text)
        key = match.group(1) if match else reading.name.lower()
        dimms[key] = max(reading.value, dimms.get(key, -100.0))

    if not dimms:
        return 0.0

    memory_type = ram_memory_type_label(temperatures)
    base_watts = {"DDR5": 4.5, "DDR4": 3.2, "LPDDR": 2.0}.get(memory_type, 3.6)
    watts = 0.0
    for temp_c in dimms.values():
        watts += base_watts
        if temp_c >= 50.0:
            watts += 1.0
        if temp_c >= 60.0:
            watts += 1.0
    return watts


def estimate_ram_label(temperatures: list[SensorReading]) -> str:
    memory_type = ram_memory_type_label(temperatures)
    return f"Estimated {memory_type} DIMM power"


def estimate_storage_watts(temperatures: list[SensorReading]) -> float:
    drives: dict[str, float] = {}
    for reading in temperatures:
        text = f"{reading.name} {reading.identifier}".lower()
        if "/nvme/" not in text and "nvme" not in text and "storage" not in text:
            continue
        match = re.search(r"/(?:nvme|storage)/([^/]+)", text)
        key = match.group(1) if match else reading.name.lower()
        drives[key] = max(reading.value, drives.get(key, -100.0))

    watts = 0.0
    for temp_c in drives.values():
        watts += 4.0
        if temp_c >= 50.0:
            watts += 1.0
        if temp_c >= 65.0:
            watts += 2.0
    return watts


def annotate_pascal_gpu_hotspots(temperatures: list[SensorReading]) -> list[str]:
    core_by_gpu: dict[str, SensorReading] = {}
    for reading in temperatures:
        group = gpu_temperature_group_key(reading)
        if not group or not is_pascal_nvidia_temperature(reading):
            continue
        if is_gpu_core_temperature(reading):
            core_by_gpu[group] = reading

    annotated = False
    for reading in temperatures:
        group = gpu_temperature_group_key(reading)
        if not group or not is_pascal_nvidia_temperature(reading) or not is_gpu_hotspot_temperature(reading):
            continue

        core = core_by_gpu.get(group)
        if core is None:
            continue

        delta = reading.value - core.value
        if 10.0 <= delta <= 20.0:
            reading.max_temperature_eligible = False
            reading.note = f"Pascal offset +{delta:.0f} C"
            annotated = True

    if not annotated:
        return []
    return [
        "Pascal NVIDIA GPU hot spot is shown as an offset reading and excluded from max temperature; GPU core is used for that card."
    ]


def gpu_temperature_group_key(reading: SensorReading) -> str:
    identifier = reading.identifier.lower().strip()
    path_parts = [part for part in identifier.strip("/").split("/") if part]
    if len(path_parts) >= 2 and path_parts[0].startswith("gpu"):
        return f"{path_parts[0]}:{path_parts[1]}"

    name = reading.name.lower()
    for marker in (" gpu hot spot", " gpu hotspot", " gpu core", " hot spot", " hotspot"):
        if marker in name:
            return name.split(marker, 1)[0].strip()
    return ""


def is_pascal_nvidia_temperature(reading: SensorReading) -> bool:
    text = f"{reading.name} {reading.identifier} {reading.source}".lower()
    if not any(token in text for token in ("nvidia", "geforce", "gtx", "quadro", "tesla", "titan")):
        return False
    return any(
        (
            re.search(r"\bgtx\s*10(50|60|70|80)(?:\s*ti)?\b", text),
            "titan xp" in text,
            "titan x pascal" in text,
            re.search(r"\bquadro\s+p\d{3,4}\b", text),
            re.search(r"\btesla\s+p(4|40|100)\b", text),
        )
    )


def is_gpu_hotspot_temperature(reading: SensorReading) -> bool:
    text = f"{reading.name} {reading.identifier}".lower()
    return "hot spot" in text or "hotspot" in text


def is_gpu_core_temperature(reading: SensorReading) -> bool:
    text = f"{reading.name} {reading.identifier}".lower()
    return "gpu core" in text and not is_gpu_hotspot_temperature(reading)


def power_group_key(reading: SensorReading) -> str:
    identifier = reading.identifier.lower().strip()
    text = f"{identifier} {reading.name} {reading.source}".lower()

    if identifier.startswith("nvidia-smi:gpu:"):
        parts = identifier.split(":")
        index = parts[2] if len(parts) > 2 else "0"
        return f"gpu:nvidia:{index}"

    path_parts = [part for part in identifier.strip("/").split("/") if part]
    if len(path_parts) >= 2:
        hardware, index = path_parts[0], path_parts[1]
        if "nvidia" in hardware:
            return f"gpu:nvidia:{index}"
        if "ati" in hardware or "amd" in hardware and "gpu" in hardware:
            return f"gpu:amd:{index}"
        if "gpu" in hardware:
            return f"gpu:{hardware}:{index}"
        if "cpu" in hardware or "processor" in hardware:
            return f"cpu:{index}"
        if "storage" in hardware or "hdd" in hardware or "ssd" in hardware:
            return f"storage:{index}"
        if "battery" in hardware:
            return f"battery:{index}"
        return f"{hardware}:{index}"

    if identifier.startswith("estimate:cpu"):
        return "cpu:estimate"
    if identifier.startswith("estimate:system"):
        return "system:estimate"
    if "cpu" in text or "package" in text or "processor" in text:
        return "cpu:0"
    if "gpu" in text or "nvidia" in text or "radeon" in text or "graphics" in text:
        return "gpu:0"
    if "battery" in text:
        return "battery:0"
    return f"misc:{reading.name.lower()}"


def component_label(reading: SensorReading) -> str:
    text = f"{reading.identifier} {reading.name} {reading.source}".lower()
    if reading.estimated:
        return "Estimate"
    if "nvidia-smi" in text or "gpu" in text or "nvidia" in text or "radeon" in text:
        return "GPU"
    if "nvme" in text or "m.2" in text:
        return "M.2 / NVMe"
    if "windows-storage" in text or "storage" in text or "ssd" in text or "hdd" in text or "disk" in text:
        return "Storage"
    if "memory" in text or "ram" in text or "dram" in text or "dimm" in text or "ddr" in text:
        return "RAM"
    if "cpu" in text or "processor" in text or "package" in text:
        return "CPU"
    if "acpi" in text or "lpc" in text or "motherboard" in text or "chipset" in text or "pch" in text:
        return "Motherboard"
    return "System"


def source_summary(readings: list[SensorReading]) -> str:
    sources = sorted({reading.source for reading in readings if reading.source})
    return ", ".join(sources) if sources else "none"


def display_source(reading: SensorReading) -> str:
    if reading.note:
        return f"{reading.source} ({reading.note})"
    return reading.source


def short_identifier(reading: SensorReading) -> str:
    identifier = reading.identifier.strip()
    if not identifier:
        return ""
    if len(identifier) <= 90:
        return identifier
    return f"...{identifier[-87:]}"


def power_selection_score(reading: SensorReading) -> float:
    text = f"{reading.name} {reading.identifier} {reading.source}".lower()
    score = 0.0

    if reading.estimated:
        score -= 40.0
    if "total" in text:
        score += 35.0
    if "package" in text:
        score += 30.0
    if "board power" in text or "total graphics power" in text:
        score += 30.0
    if "gpu power" in text or "power draw" in text:
        score += 18.0
    if "cpu package" in text:
        score += 20.0
    if "nvidia-smi" in text:
        score += 8.0
    if "core" in text and "gpu core" not in text:
        score -= 12.0
    if "dram" in text or "uncore" in text or "gt cores" in text:
        score -= 8.0
    score += clamp(reading.value, 0.0, 500.0) / 1000.0
    return score


def select_power_for_total(power: list[SensorReading]) -> list[SensorReading]:
    grouped: dict[str, list[SensorReading]] = defaultdict(list)
    for reading in power:
        grouped[power_group_key(reading)].append(reading)

    selected: list[SensorReading] = []
    for group, readings in grouped.items():
        if group.startswith("misc:") and len(readings) > 1:
            selected.extend(readings)
            continue
        best = max(readings, key=power_selection_score)
        if not best.estimated and best.value <= 0.5:
            continue
        best.selected_for_total = True
        selected.append(best)
    return selected


class Poller(threading.Thread):
    def __init__(self, output_queue: queue.Queue[HardwareSnapshot], stop_event: threading.Event) -> None:
        super().__init__(daemon=True)
        self.collector = SensorCollector()
        self.output_queue = output_queue
        self.stop_event = stop_event

    def run(self) -> None:
        while not self.stop_event.is_set():
            snapshot = self.collector.sample()
            self.output_queue.put(snapshot)
            self.stop_event.wait(POLL_INTERVAL_SECONDS)


class StatCard(tk.Frame):
    def __init__(self, master: tk.Widget, title: str, accent: str) -> None:
        super().__init__(
            master,
            bg=COLORS["card"],
            highlightbackground=COLORS["border"],
            highlightthickness=1,
            bd=0,
        )
        self.accent = accent
        self.grid_columnconfigure(0, weight=1)
        tk.Frame(self, bg=accent, width=4).grid(row=0, column=0, sticky="nsw", rowspan=3)
        self.title_label = tk.Label(
            self,
            text=title.upper(),
            bg=COLORS["card"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9, "bold"),
            anchor="w",
        )
        self.title_label.grid(row=0, column=0, sticky="ew", padx=(16, 12), pady=(12, 0))
        self.value_label = tk.Label(
            self,
            text="--",
            bg=COLORS["card"],
            fg=COLORS["text"],
            font=("Segoe UI", 22, "bold"),
            anchor="w",
        )
        self.value_label.grid(row=1, column=0, sticky="ew", padx=(16, 12), pady=(2, 0))
        self.sub_label = tk.Label(
            self,
            text="Waiting for data",
            bg=COLORS["card"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9),
            anchor="w",
        )
        self.sub_label.grid(row=2, column=0, sticky="ew", padx=(16, 12), pady=(0, 12))

    def set(self, value: str, subtext: str) -> None:
        self.value_label.configure(text=value)
        self.sub_label.configure(text=subtext)


class GraphPanel(tk.Frame):
    def __init__(self, master: tk.Widget) -> None:
        super().__init__(
            master,
            bg=COLORS["panel"],
            highlightbackground=COLORS["border"],
            highlightthickness=1,
            bd=0,
        )
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        tk.Label(
            self,
            text="Trends",
            bg=COLORS["panel"],
            fg=COLORS["text"],
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="ew", padx=14, pady=(10, 0))
        self.canvas = tk.Canvas(self, bg=COLORS["panel"], highlightthickness=0, height=248)
        self.canvas.grid(row=1, column=0, sticky="nsew", padx=10, pady=(4, 10))
        self.canvas.bind("<Configure>", lambda _event: self.redraw())
        self.history: deque[dict[str, object]] = deque(maxlen=2000)
        self.show_time = True
        self.x_window_seconds: Optional[float] = None
        self.y_scale_mode = "auto"

    def apply_options(
        self,
        *,
        show_time: bool,
        x_window_seconds: Optional[float],
        y_scale_mode: str,
    ) -> None:
        self.show_time = show_time
        self.x_window_seconds = x_window_seconds
        self.y_scale_mode = y_scale_mode
        self.redraw()

    def set_show_time(self, enabled: bool) -> None:
        self.show_time = enabled
        self.redraw()

    def _visible_samples(self) -> list[dict[str, object]]:
        if not self.history:
            return []
        samples = list(self.history)
        if self.x_window_seconds is None:
            return samples
        cutoff = time.time() - self.x_window_seconds
        return [sample for sample in samples if float(sample["ts"]) >= cutoff]

    def _value_bounds(self, values: list[float], key: str, unit: str) -> tuple[float, float]:
        low = min(values)
        high = max(values)
        if math.isclose(low, high):
            padding = 1.0 if unit == "C" else max(1.0, abs(high) * 0.05)
            low = low - padding if key == "temp" else max(0.0, low - padding)
            high += padding

        if self.y_scale_mode == "zero" and key in ("watts", "btu"):
            low = 0.0
            if high <= 0.0:
                high = 10.0
        elif self.y_scale_mode == "padded":
            span = high - low
            padding = span * 0.10 if span > 0 else (1.0 if unit == "C" else max(1.0, abs(high) * 0.1))
            low -= padding
            high += padding
            if key in ("watts", "btu"):
                low = max(0.0, low)
        elif self.y_scale_mode == "zero" and key == "temp":
            span = high - low
            padding = max(1.0, span * 0.08)
            low -= padding
            high += padding

        return low, high

    def add_sample(
        self,
        watts: float,
        btu_per_hour: float,
        max_temp: Optional[float],
        timestamp: Optional[float] = None,
    ) -> None:
        self.history.append(
            {
                "watts": watts,
                "btu": btu_per_hour,
                "temp": max_temp,
                "ts": timestamp if timestamp is not None else time.time(),
            }
        )
        self.redraw()

    def redraw(self) -> None:
        canvas = self.canvas
        canvas.delete("all")
        width = max(canvas.winfo_width(), 300)
        height = max(canvas.winfo_height(), 180)
        pad_x = 72
        right_pad = 16
        top_pad = 8
        band_gap = 10
        time_axis_h = 40 if self.show_time else 0
        plot_height = height - top_pad - time_axis_h
        band_height = (plot_height - band_gap * 2) / 3.0

        bands = [
            ("watts", "Watts", "W", COLORS["green"]),
            ("btu", "BTU/hr", "BTU/hr", COLORS["amber"]),
            ("temp", "Max temp", "C", COLORS["coral"]),
        ]

        for index, (key, label, unit, color) in enumerate(bands):
            y = top_pad + index * (band_height + band_gap)
            self._draw_band(canvas, key, label, unit, color, pad_x, y, width - right_pad, band_height)

        if self.show_time:
            self._draw_time_axis(canvas, pad_x, width - right_pad, top_pad + plot_height, time_axis_h)

    def _draw_time_axis(
        self,
        canvas: tk.Canvas,
        left: float,
        right: float,
        top: float,
        axis_height: float,
    ) -> None:
        line_y = top + 10
        label_y = top + axis_height - 6
        canvas.create_line(left, line_y, right, line_y, fill=COLORS["grid"])

        if not self.history:
            canvas.create_text(
                (left + right) / 2,
                label_y,
                text="Time",
                fill=COLORS["muted"],
                font=("Segoe UI", 9),
                anchor="s",
            )
            return

        visible = self._visible_samples()
        timestamps = [float(sample["ts"]) for sample in visible]
        if not timestamps:
            return
        span = timestamps[-1] - timestamps[0] if len(timestamps) > 1 else 0.0
        include_seconds = span < 120.0
        plot_width = right - left
        tick_count = 3 if plot_width < 420 else 5 if plot_width < 700 else 7
        tick_count = min(tick_count, len(timestamps))
        if tick_count < 2:
            tick_count = len(timestamps)

        if tick_count == 1:
            positions = [0.0]
        else:
            positions = [index / (tick_count - 1) for index in range(tick_count)]

        seen_labels: set[str] = set()
        for position in positions:
            index = int(round(position * (len(timestamps) - 1)))
            index = clamp(index, 0, len(timestamps) - 1)
            x = left + position * (right - left)
            label = format_graph_time(timestamps[index], include_seconds=include_seconds)
            if label in seen_labels:
                continue
            seen_labels.add(label)
            canvas.create_line(x, line_y - 5, x, line_y + 5, fill=COLORS["muted"])
            canvas.create_text(
                x,
                label_y,
                text=label,
                fill=COLORS["text"],
                font=("Segoe UI", 9),
                anchor="s",
            )

    def _draw_band(
        self,
        canvas: tk.Canvas,
        key: str,
        label: str,
        unit: str,
        color: str,
        left: float,
        top: float,
        right: float,
        height: float,
    ) -> None:
        bottom = top + height
        canvas.create_rectangle(left, top, right, bottom, fill=COLORS["bg"], outline=COLORS["grid"])

        visible = self._visible_samples()
        values = [float(sample[key]) for sample in visible if sample.get(key) is not None]
        current = values[-1] if values else None
        current_text = self._format_axis_value(current, unit)
        canvas.create_text(
            8,
            top + 14,
            text=label,
            fill=COLORS["muted"],
            font=("Segoe UI", 9, "bold"),
            anchor="w",
        )
        canvas.create_text(
            8,
            top + 34,
            text=current_text,
            fill=color,
            font=("Segoe UI", 10, "bold"),
            anchor="w",
        )

        if len(values) < 2:
            canvas.create_text(
                (left + right) / 2,
                (top + bottom) / 2,
                text="Collecting samples",
                fill=COLORS["subtle"],
                font=("Segoe UI", 10),
            )
            return

        low, high = self._value_bounds(values, key, unit)

        plot_top = top + 5
        plot_bottom = bottom - 5
        plot_height = plot_bottom - plot_top

        for fraction in (1.0 / 3.0, 2.0 / 3.0):
            grid_y = plot_top + fraction * plot_height
            canvas.create_line(left, grid_y, right, grid_y, fill=COLORS["grid"])

        for fraction in (0.0, 1.0 / 3.0, 2.0 / 3.0, 1.0):
            grid_value = high - fraction * (high - low)
            grid_y = plot_top + fraction * plot_height
            canvas.create_text(
                right - 6,
                grid_y,
                text=self._format_grid_value(grid_value, unit),
                fill=COLORS["muted"],
                font=("Segoe UI", 8),
                anchor="e",
            )

        samples = [sample.get(key) for sample in visible]
        x_step = (right - left) / max(1, len(samples) - 1)
        points: list[float] = []
        last_valid: Optional[tuple[float, float]] = None
        for index, raw_value in enumerate(samples):
            if raw_value is None:
                continue
            value = float(raw_value)
            x = left + index * x_step
            normalized = (value - low) / (high - low)
            y = bottom - normalized * (height - 10) - 5
            if last_valid is not None and points:
                points.extend([x, y])
            else:
                points = [x, y]
            last_valid = (x, y)

        if len(points) >= 4:
            canvas.create_line(points, fill=color, width=2.4, smooth=True)

    def _format_grid_value(self, value: float, unit: str) -> str:
        if unit == "C":
            return f"{value:.1f}"
        return f"{value:.0f}"

    def _format_axis_value(self, value: Optional[float], unit: str) -> str:
        if value is None:
            return "--"
        if unit == "BTU/hr":
            return f"{value:.0f}"
        if unit == "W":
            return f"{value:.0f} W"
        return f"{value:.0f} C"


class MetricTable(tk.Frame):
    def __init__(self, master: tk.Widget, title: str, columns: tuple[str, ...]) -> None:
        super().__init__(
            master,
            bg=COLORS["panel"],
            highlightbackground=COLORS["border"],
            highlightthickness=1,
            bd=0,
        )
        self.columns = columns
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        tk.Label(
            self,
            text=title,
            bg=COLORS["panel"],
            fg=COLORS["text"],
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="ew", padx=14, pady=(10, 6))

        self.tree = ttk.Treeview(
            self,
            columns=columns,
            show="headings",
            selectmode="none",
            height=7,
        )
        self.tree.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        for column in columns:
            self.tree.heading(column, text=column)
        self.tree.column(columns[0], width=230, stretch=True, anchor="w")
        for column in columns[1:]:
            self.tree.column(column, width=110, stretch=False, anchor="w")

    def set_rows(self, rows: list[tuple[str, ...]]) -> None:
        self.tree.delete(*self.tree.get_children())
        if not rows:
            self.tree.insert("", "end", values=tuple(["No readings yet"] + [""] * (len(self.columns) - 1)))
            return
        for row in rows:
            self.tree.insert("", "end", values=row)


class OptionsWindow(tk.Toplevel):
    def __init__(self, master: tk.Tk, app: "HeatLensApp") -> None:
        super().__init__(master)
        self.app = app
        self.title(f"{APP_NAME} Options")
        self.configure(bg=COLORS["bg"])
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        frame = tk.Frame(self, bg=COLORS["bg"], padx=18, pady=16)
        frame.grid(row=0, column=0, sticky="nsew")
        frame.grid_columnconfigure(1, weight=1)

        row = 0

        tk.Label(
            frame,
            text="Trend graphs",
            bg=COLORS["bg"],
            fg=COLORS["text"],
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        ).grid(row=row, column=0, columnspan=2, sticky="w", pady=(0, 10))
        row += 1

        ttk.Checkbutton(
            frame,
            text="Show time on X-axis",
            variable=app.show_graph_time,
            command=app._apply_graph_options,
            style="HeatLens.TCheckbutton",
        ).grid(row=row, column=0, columnspan=2, sticky="w")
        row += 1

        tk.Label(
            frame,
            text="X-axis window",
            bg=COLORS["bg"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=row, column=0, sticky="w", pady=(10, 4))
        self.x_scale_combo = ttk.Combobox(
            frame,
            textvariable=app.graph_x_window,
            values=GRAPH_X_WINDOW_LABELS,
            state="readonly",
            width=18,
        )
        self.x_scale_combo.grid(row=row, column=1, sticky="e", pady=(10, 4))
        self.x_scale_combo.bind("<<ComboboxSelected>>", lambda _event: app._apply_graph_options())
        row += 1

        tk.Label(
            frame,
            text="Y-axis scaling",
            bg=COLORS["bg"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=row, column=0, sticky="w", pady=(4, 4))
        self.y_scale_combo = ttk.Combobox(
            frame,
            textvariable=app.graph_y_scale,
            values=GRAPH_Y_SCALE_LABELS,
            state="readonly",
            width=18,
        )
        self.y_scale_combo.grid(row=row, column=1, sticky="e", pady=(4, 4))
        self.y_scale_combo.bind("<<ComboboxSelected>>", lambda _event: app._apply_graph_options())
        row += 1

        tk.Label(
            frame,
            text="Include zero applies to watts and BTU/hr. Temperature always auto-scales to the visible range.",
            bg=COLORS["bg"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9),
            anchor="w",
            wraplength=360,
            justify="left",
        ).grid(row=row, column=0, columnspan=2, sticky="w", pady=(6, 14))
        row += 1

        if sys.platform == "win32":
            tk.Label(
                frame,
                text="Windows",
                bg=COLORS["bg"],
                fg=COLORS["text"],
                font=("Segoe UI", 12, "bold"),
                anchor="w",
            ).grid(row=row, column=0, columnspan=2, sticky="w", pady=(0, 10))
            row += 1

            admin_text = (
                "Running as administrator."
                if is_windows_admin()
                else "Restart as administrator for extra ACPI, storage, and WMI sensors."
            )
            self.admin_label = tk.Label(
                frame,
                text=admin_text,
                bg=COLORS["bg"],
                fg=COLORS["muted"],
                font=("Segoe UI", 9),
                anchor="w",
                wraplength=360,
                justify="left",
            )
            self.admin_label.grid(row=row, column=0, columnspan=2, sticky="w", pady=(0, 8))
            row += 1

            ttk.Checkbutton(
                frame,
                text="Always start as administrator",
                variable=app.always_start_as_admin,
                command=app._apply_always_start_as_admin,
                style="HeatLens.TCheckbutton",
            ).grid(row=row, column=0, columnspan=2, sticky="w", pady=(0, 8))
            row += 1

            self.admin_button = ttk.Button(
                frame,
                text="Restart as administrator",
                command=app._restart_as_admin,
                style="HeatLens.TButton",
                state="disabled" if is_windows_admin() else "normal",
            )
            self.admin_button.grid(row=row, column=0, columnspan=2, sticky="w", pady=(0, 14))
            row += 1

        buttons = tk.Frame(frame, bg=COLORS["bg"])
        buttons.grid(row=row, column=0, columnspan=2, sticky="e")
        ttk.Button(
            buttons,
            text="Close",
            command=self.destroy,
            style="HeatLens.TButton",
        ).grid(row=0, column=0)

        self.update_idletasks()
        x = master.winfo_rootx() + max(0, (master.winfo_width() - self.winfo_width()) // 2)
        y = master.winfo_rooty() + max(0, (master.winfo_height() - self.winfo_height()) // 2)
        self.geometry(f"+{x}+{y}")


class SensorSourceWindow(tk.Toplevel):
    def __init__(self, master: tk.Widget) -> None:
        super().__init__(master)
        self.title("HeatLens Sensors")
        self.configure(bg=COLORS["bg"])
        self.geometry("1040x520")
        self.minsize(780, 420)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.summary_label = tk.Label(
            self,
            text="Waiting for sensor scan",
            bg=COLORS["bg"],
            fg=COLORS["text"],
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        )
        self.summary_label.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))

        table_frame = tk.Frame(self, bg=COLORS["panel"])
        table_frame.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 10))
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        columns = ("Type", "Component", "Sensor", "Value", "Source", "Used", "Identifier")
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
            height=13,
        )
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        widths = {
            "Type": 80,
            "Component": 115,
            "Sensor": 285,
            "Value": 130,
            "Source": 155,
            "Used": 80,
            "Identifier": 340,
        }
        for column in columns:
            self.tree.heading(column, text=column)
            self.tree.column(column, width=widths[column], stretch=column in {"Sensor", "Identifier"}, anchor="w")

        self.note_label = tk.Label(
            self,
            text="",
            bg=COLORS["bg"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9),
            anchor="w",
            justify="left",
        )
        self.note_label.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 12))

    def update_snapshot(self, snapshot: HardwareSnapshot) -> None:
        self.tree.delete(*self.tree.get_children())
        readings = list(snapshot.power) + list(snapshot.temperatures)
        sources = source_summary(readings)
        self.summary_label.configure(
            text=(
                f"{len(snapshot.power)} power readings | "
                f"{len(snapshot.temperatures)} temperature readings | Sources: {sources}"
            )
        )

        for reading in snapshot.power:
            self.tree.insert(
                "",
                "end",
                values=(
                    "Power",
                    component_label(reading),
                    reading.name,
                    reading.display_value(),
                    display_source(reading),
                    self._used_label(reading),
                    short_identifier(reading),
                ),
            )
        for reading in snapshot.temperatures:
            self.tree.insert(
                "",
                "end",
                values=(
                    "Temp",
                    component_label(reading),
                    reading.name,
                    reading.display_value(),
                    display_source(reading),
                    "",
                    short_identifier(reading),
                ),
            )

        if not readings:
            self.tree.insert("", "end", values=("None", "", "No sensors found yet", "", "", "", ""))

        if snapshot.notes:
            self.note_label.configure(text=" | ".join(snapshot.notes))
        else:
            self.note_label.configure(text="All visible sensors are listed here. Total wattage uses rows marked Total or Estimate.")

    def _used_label(self, reading: SensorReading) -> str:
        if reading.estimated and reading.selected_for_total:
            return "Estimate"
        if reading.selected_for_total:
            return "Total"
        return ""


class HeatLensApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(f"{APP_NAME} {APP_VERSION}")
        self.root.configure(bg=COLORS["bg"])
        self.root.geometry("1060x720")
        self.root.minsize(820, 600)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        self.queue: queue.Queue[HardwareSnapshot] = queue.Queue()
        self.stop_event = threading.Event()
        self.poller = Poller(self.queue, self.stop_event)

        self.last_snapshot: Optional[HardwareSnapshot] = None
        self.session_wh = 0.0
        self.session_elapsed_seconds = 0.0
        self.session_log: list[SessionLogEntry] = []
        self.started_at = time.time()
        self.compact = False
        self.details_visible = True
        self.source_window: Optional[SensorSourceWindow] = None
        self.options_window: Optional[OptionsWindow] = None
        self.always_on_top = tk.BooleanVar(value=False)
        self.show_graph_time = tk.BooleanVar(value=True)
        self.graph_x_window = tk.StringVar(value="Auto")
        self.graph_y_scale = tk.StringVar(value="Auto")
        self.ambient_temp_var = tk.StringVar(value=f"{DEFAULT_AMBIENT_TEMP_F:.0f}")
        self.preferences = HeatLensPreferences()
        self.always_start_as_admin = tk.BooleanVar(
            value=self.preferences.get_bool(PREF_ALWAYS_START_AS_ADMIN, False)
        )
        self.libre_helper = LibreHardwareMonitorHelper()

        configure_ttk_style()
        self._build_ui()
        self._apply_graph_options()
        self.poller.start()
        self.root.after(150, self._drain_queue)
        self.root.after(900, lambda: self.libre_helper.maybe_prompt_on_startup(
            self.root,
            on_status=self._libre_status,
            on_connected=self._libre_connected,
            on_failed=self._libre_wait_failed,
        ))

    def _build_ui(self) -> None:
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(2, weight=1)

        header = tk.Frame(self.root, bg=COLORS["bg"])
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 10))
        header.grid_columnconfigure(1, weight=1)

        title_block = tk.Frame(header, bg=COLORS["bg"])
        title_block.grid(row=0, column=0, sticky="w")
        tk.Label(
            title_block,
            text=APP_NAME,
            bg=COLORS["bg"],
            fg=COLORS["text"],
            font=("Segoe UI", 20, "bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        self.status_label = tk.Label(
            title_block,
            text="Starting sensor scan",
            bg=COLORS["bg"],
            fg=COLORS["muted"],
            font=("Segoe UI", 10),
            anchor="w",
        )
        self.status_label.grid(row=1, column=0, sticky="w")

        controls = tk.Frame(header, bg=COLORS["bg"])
        controls.grid(row=0, column=2, sticky="e")
        tk.Label(
            controls,
            text="Ambient",
            bg=COLORS["bg"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=0, column=0, padx=(0, 6))
        self.ambient_entry = ttk.Entry(
            controls,
            width=6,
            textvariable=self.ambient_temp_var,
            style="HeatLens.TEntry",
        )
        self.ambient_entry.grid(row=0, column=1, padx=(0, 4))
        self.ambient_entry.bind("<Return>", self._refresh_current_snapshot)
        self.ambient_entry.bind("<FocusOut>", self._refresh_current_snapshot)
        tk.Label(
            controls,
            text="F/C",
            bg=COLORS["bg"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=0, column=2, padx=(0, 10))
        self.top_check = ttk.Checkbutton(
            controls,
            text="Pin",
            variable=self.always_on_top,
            command=self._toggle_always_on_top,
            style="HeatLens.TCheckbutton",
        )
        self.top_check.grid(row=0, column=3, padx=(0, 8))
        self.libre_button = ttk.Button(
            controls,
            text="Libre",
            command=self._setup_libre_hardware_monitor,
            style="HeatLens.TButton",
        )
        self.libre_button.grid(row=0, column=4, padx=(0, 8))
        self.sources_button = ttk.Button(
            controls,
            text="Sensors",
            command=self._show_sources,
            style="HeatLens.TButton",
        )
        self.sources_button.grid(row=0, column=5, padx=(0, 8))
        self.export_button = ttk.Button(
            controls,
            text="Export",
            command=self._export_log,
            style="HeatLens.TButton",
        )
        self.export_button.grid(row=0, column=6, padx=(0, 8))
        self.compact_button = ttk.Button(
            controls,
            text="Compact",
            command=self._toggle_compact,
            style="HeatLens.TButton",
        )
        self.compact_button.grid(row=0, column=7, padx=(0, 8))
        self.options_button = ttk.Button(
            controls,
            text="Options",
            command=self._show_options,
            style="HeatLens.TButton",
        )
        self.options_button.grid(row=0, column=8)

        stats = tk.Frame(self.root, bg=COLORS["bg"])
        stats.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))
        for col in range(4):
            stats.grid_columnconfigure(col, weight=1, uniform="stats")

        self.watts_card = StatCard(stats, "Total Wattage", COLORS["green"])
        self.watts_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self.heat_card = StatCard(stats, "Heat Dissipation", COLORS["amber"])
        self.heat_card.grid(row=0, column=1, sticky="nsew", padx=8)
        self.session_card = StatCard(stats, "Session Heat", COLORS["cyan"])
        self.session_card.grid(row=0, column=2, sticky="nsew", padx=8)
        self.temp_card = StatCard(stats, "Max Temperature", COLORS["coral"])
        self.temp_card.grid(row=0, column=3, sticky="nsew", padx=(8, 0))

        self.main_area = tk.Frame(self.root, bg=COLORS["bg"])
        self.main_area.grid(row=2, column=0, sticky="nsew", padx=18, pady=(0, 10))
        self.main_area.grid_columnconfigure(0, weight=1)
        self.main_area.grid_rowconfigure(0, weight=0)
        self.main_area.grid_rowconfigure(1, weight=1)

        self.graph = GraphPanel(self.main_area)
        self.graph.grid(row=0, column=0, sticky="ew")

        self.details = tk.Frame(self.main_area, bg=COLORS["bg"])
        self.details.grid(row=1, column=0, sticky="nsew", pady=(12, 0))
        self.details.grid_columnconfigure(0, weight=1)
        self.details.grid_columnconfigure(1, weight=1)
        self.details.grid_rowconfigure(0, weight=1)

        self.power_table = MetricTable(self.details, "Power Metrics", ("Metric", "Value", "Used", "Source"))
        self.power_table.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        self.temp_table = MetricTable(self.details, "Temperature Metrics", ("Sensor", "Reading", "Source"))
        self.temp_table.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        footer = tk.Frame(self.root, bg=COLORS["bg"])
        footer.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 14))
        footer.grid_columnconfigure(0, weight=1)

        self.note_label = tk.Label(
            footer,
            text="BTU/hr = watts x 3.412. For true whole-room heat, a wall power meter is the gold standard.",
            bg=COLORS["bg"],
            fg=COLORS["muted"],
            font=("Segoe UI", 9),
            anchor="w",
        )
        self.note_label.grid(row=0, column=0, sticky="w")

        if DONATE_URL.strip():
            self.donate_label = tk.Label(
                footer,
                text="Buy me a coffee",
                bg=COLORS["bg"],
                fg=COLORS["subtle"],
                font=("Segoe UI", 9, "underline"),
                cursor="hand2",
                anchor="e",
            )
            self.donate_label.grid(row=0, column=1, sticky="e", padx=(12, 0))
            self.donate_label.bind(
                "<Button-1>",
                lambda _event: self.libre_helper.open_donate_page(),
            )
            self.donate_label.bind(
                "<Enter>",
                lambda _event: self.donate_label.configure(fg=COLORS["muted"]),
            )
            self.donate_label.bind(
                "<Leave>",
                lambda _event: self.donate_label.configure(fg=COLORS["subtle"]),
            )

    def _setup_libre_hardware_monitor(self) -> None:
        self.libre_helper.prompt_manual_action(
            self.root,
            on_status=self._libre_status,
            on_connected=self._libre_connected,
            on_failed=self._libre_wait_failed,
        )

    def _libre_status(self, message: str) -> None:
        self.status_label.configure(text=message)

    def _libre_connected(self) -> None:
        self.status_label.configure(text="Connected to LibreHardwareMonitor sensors")
        messagebox.showinfo(
            "LibreHardwareMonitor",
            "Sensor feed connected. HeatLens should refresh on the next sample.",
            parent=self.root,
        )

    def _libre_wait_failed(self) -> None:
        install_path = self.libre_helper.find_installation()
        if install_path is not None and self.libre_helper.is_running():
            self.libre_helper._prompt_enable_web_server(self.root)
            self.libre_helper.begin_sensor_feed_wait(
                self.root,
                install_path,
                on_status=self._libre_status,
                on_connected=self._libre_connected,
                on_failed=lambda: self.status_label.configure(
                    text="Libre running — enable Options -> Remote Web Server -> Run"
                ),
                launch_if_needed=False,
                max_attempts=15,
            )
            return

        self.status_label.configure(text="LibreHardwareMonitor sensors not connected yet")
        if install_path is None:
            return
        if messagebox.askyesno(
            "LibreHardwareMonitor",
            "HeatLens still cannot read LibreHardwareMonitor sensors.\n\n"
            "Try launching it as administrator?",
            parent=self.root,
        ):
            if self.libre_helper.launch_elevated(install_path):
                self.libre_helper.begin_sensor_feed_wait(
                    self.root,
                    install_path,
                    on_status=self._libre_status,
                    on_connected=self._libre_connected,
                    on_failed=lambda: self.status_label.configure(
                        text="Libre admin launch done — enable Remote Web Server -> Run"
                    ),
                    launch_if_needed=False,
                )

    def _toggle_always_on_top(self) -> None:
        self.root.attributes("-topmost", self.always_on_top.get())

    def _apply_graph_options(self) -> None:
        x_label = self.graph_x_window.get()
        y_label = self.graph_y_scale.get()
        x_seconds = GRAPH_X_WINDOW_SECONDS.get(x_label, None)
        y_mode = GRAPH_Y_SCALE_MODES.get(y_label, "auto")
        self.graph.apply_options(
            show_time=self.show_graph_time.get(),
            x_window_seconds=x_seconds,
            y_scale_mode=y_mode,
        )

    def _restart_as_admin(self) -> None:
        if sys.platform != "win32":
            messagebox.showinfo("HeatLens", "Administrator mode is only available on Windows.", parent=self.root)
            return
        if is_windows_admin():
            messagebox.showinfo("HeatLens", "HeatLens is already running as administrator.", parent=self.root)
            return
        if not messagebox.askyesno(
            "Restart as administrator",
            "Windows will ask for permission, then HeatLens will restart with administrator rights.\n\n"
            "This can unlock extra ACPI, storage, and WMI sensors.\n\n"
            "Continue?",
            parent=self.root,
        ):
            return
        if not launch_heatlens_elevated():
            messagebox.showerror(
                "HeatLens",
                "Could not restart as administrator. Windows may have cancelled the prompt.",
                parent=self.root,
            )
            return
        self.on_close()

    def _apply_always_start_as_admin(self) -> None:
        enabled = self.always_start_as_admin.get()
        self.preferences.set_bool(PREF_ALWAYS_START_AS_ADMIN, enabled)
        if not enabled or sys.platform != "win32" or is_windows_admin():
            return
        if messagebox.askyesno(
            "Restart as administrator",
            "Always start as administrator is now enabled.\n\n"
            "Restart HeatLens now with administrator rights?",
            parent=self.root,
        ):
            if launch_heatlens_elevated():
                self.on_close()
            else:
                messagebox.showerror(
                    "HeatLens",
                    "Could not restart as administrator. Windows may have cancelled the prompt.",
                    parent=self.root,
                )

    def _show_options(self) -> None:
        if self.options_window is not None and self.options_window.winfo_exists():
            self.options_window.lift()
            self.options_window.focus_force()
            return
        self.options_window = OptionsWindow(self.root, self)
        self.options_window.protocol("WM_DELETE_WINDOW", self._close_options)

    def _close_options(self) -> None:
        if self.options_window is not None and self.options_window.winfo_exists():
            self.options_window.destroy()
        self.options_window = None

    def _show_sources(self) -> None:
        if self.source_window is not None and self.source_window.winfo_exists():
            self.source_window.lift()
            self.source_window.focus_force()
        else:
            self.source_window = SensorSourceWindow(self.root)
            self.source_window.protocol("WM_DELETE_WINDOW", self._close_sources)
        if self.last_snapshot is not None:
            self.source_window.update_snapshot(self.last_snapshot)

    def _close_sources(self) -> None:
        if self.source_window is not None and self.source_window.winfo_exists():
            self.source_window.destroy()
        self.source_window = None

    def _refresh_current_snapshot(self, _event: object | None = None) -> None:
        if self.last_snapshot is not None:
            self._apply_snapshot(self.last_snapshot, update_session=False)

    def _ambient_temp_f(self) -> Optional[float]:
        return parse_ambient_temperature_f(self.ambient_temp_var.get())

    def _export_log(self) -> None:
        if not self.session_log:
            messagebox.showinfo("HeatLens Export", "No logged samples are available yet.")
            return

        default_name = f"HeatLens_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            parent=self.root,
            title="Export HeatLens log",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=default_name,
        )
        if not path:
            return

        try:
            export_session_log_to_excel(self.session_log, path)
        except ModuleNotFoundError as exc:
            if exc.name == "openpyxl":
                messagebox.showerror(
                    "HeatLens Export",
                    "Excel export needs openpyxl. Run:\n\npy -3 -m pip install -r requirements_heat_widget.txt",
                )
                return
            messagebox.showerror("HeatLens Export", f"Export failed: {exc}")
        except Exception as exc:
            messagebox.showerror("HeatLens Export", f"Export failed: {exc}")
        else:
            messagebox.showinfo("HeatLens Export", f"Exported {len(self.session_log)} samples to:\n{path}")

    def _toggle_compact(self) -> None:
        self.compact = not self.compact
        if self.compact:
            self.details.grid_remove()
            self.root.geometry("760x430")
            self.compact_button.configure(text="Detailed")
        else:
            self.details.grid()
            self.root.geometry("1060x720")
            self.compact_button.configure(text="Compact")

    def _drain_queue(self) -> None:
        latest: Optional[HardwareSnapshot] = None
        while True:
            try:
                latest = self.queue.get_nowait()
            except queue.Empty:
                break
        if latest is not None:
            self._apply_snapshot(latest)
        if not self.stop_event.is_set():
            self.root.after(250, self._drain_queue)

    def _apply_snapshot(self, snapshot: HardwareSnapshot, update_session: bool = True) -> None:
        if update_session:
            self._update_session_energy(snapshot)
        self.last_snapshot = snapshot
        session_btu = self.session_wh * WATTS_TO_BTU_PER_HOUR
        ambient_f = self._ambient_temp_f()

        direct_count = len([reading for reading in snapshot.selected_power if not reading.estimated])
        estimate_count = len([reading for reading in snapshot.selected_power if reading.estimated])
        selected_label = f"{direct_count} sensor"
        if direct_count != 1:
            selected_label += "s"
        if estimate_count:
            selected_label += f" + {estimate_count} est"
        average_label = self._average_energy_rate_label(snapshot)
        if average_label:
            selected_label = f"{selected_label} | {average_label}"

        self.status_label.configure(text=snapshot.status)
        self.watts_card.set(format_watts(snapshot.total_watts), selected_label)
        self.heat_card.set(format_btu_per_hour(snapshot.btu_per_hour), heat_equivalent_label(snapshot.total_watts))
        self.session_card.set(format_btu(session_btu), f"Since launch, {self._elapsed_text()}")
        temp_count = len(snapshot.temperatures)
        temp_label = f"{temp_count} temperature sensor" + ("" if temp_count == 1 else "s")
        self.temp_card.set(format_temp(snapshot.max_temp_c), temp_label)

        if update_session:
            self._append_log_entry(snapshot, ambient_f)
            self.graph.add_sample(
                snapshot.total_watts,
                snapshot.btu_per_hour,
                snapshot.max_temp_c,
                timestamp=snapshot.taken_at,
            )
        self.power_table.set_rows(self._power_rows(snapshot, ambient_f))
        self.temp_table.set_rows(self._temperature_rows(snapshot, ambient_f))
        if self.source_window is not None and self.source_window.winfo_exists():
            self.source_window.update_snapshot(snapshot)
        self.note_label.configure(text=self._status_note(snapshot, ambient_f))

    def _update_session_energy(self, snapshot: HardwareSnapshot) -> None:
        if self.last_snapshot is None:
            return
        elapsed = clamp(snapshot.taken_at - self.last_snapshot.taken_at, 0.0, 10.0)
        average_watts = (snapshot.total_watts + self.last_snapshot.total_watts) / 2.0
        self.session_wh += average_watts * elapsed / 3600.0
        self.session_elapsed_seconds += elapsed

    def _elapsed_text(self) -> str:
        elapsed = max(0, int(time.time() - self.started_at))
        minutes, seconds = divmod(elapsed, 60)
        hours, minutes = divmod(minutes, 60)
        if hours:
            return f"{hours}h {minutes}m"
        if minutes:
            return f"{minutes}m {seconds}s"
        return f"{seconds}s"

    def _average_watts(self) -> Optional[float]:
        elapsed_hours = self.session_elapsed_seconds / 3600.0
        if elapsed_hours <= 0.0 or self.session_wh <= 0.0:
            return None
        return self.session_wh / elapsed_hours

    def _average_energy_rate_label(self, snapshot: HardwareSnapshot) -> str:
        average_watts = self._average_watts()
        if average_watts is None:
            return ""
        return f"avg {format_kwh_per_hour(average_watts)}"

    def _append_log_entry(self, snapshot: HardwareSnapshot, ambient_f: Optional[float]) -> None:
        direct_watts = sum(max(0.0, reading.value) for reading in snapshot.selected_power if not reading.estimated)
        estimated_watts = sum(max(0.0, reading.value) for reading in snapshot.selected_power if reading.estimated)
        selected_sources = "; ".join(
            f"{reading.name}={reading.display_value()}" for reading in snapshot.selected_power
        )
        notes = " | ".join(snapshot.notes)
        average_watts = self._average_watts()
        self.session_log.append(
            SessionLogEntry(
                timestamp=datetime.fromtimestamp(snapshot.taken_at),
                elapsed_seconds=max(0.0, snapshot.taken_at - self.started_at),
                total_watts=snapshot.total_watts,
                direct_watts=direct_watts,
                estimated_watts=estimated_watts,
                average_kwh_per_hour=None if average_watts is None else average_watts / 1000.0,
                btu_per_hour=snapshot.btu_per_hour,
                max_temp_c=snapshot.max_temp_c,
                ambient_f=ambient_f,
                heat_equivalent=heat_equivalent_label(snapshot.total_watts),
                status=snapshot.status,
                selected_sources=selected_sources,
                notes=notes,
            )
        )

    def _power_rows(self, snapshot: HardwareSnapshot, ambient_f: Optional[float]) -> list[tuple[str, ...]]:
        rows: list[tuple[str, ...]] = []
        rows.append(("Total wattage", format_watts(snapshot.total_watts), "Total", "Derived"))
        rows.append(("Heat dissipation", format_btu_per_hour(snapshot.btu_per_hour), "Derived", "Derived"))
        rows.append(("Heat equivalent", heat_equivalent_label(snapshot.total_watts), "20-level", "Approximation"))
        if ambient_f is not None:
            air_rise = still_air_rise_f_per_hour(snapshot.btu_per_hour, ambient_f)
            cfm = airflow_for_exhaust_rise_cfm(snapshot.btu_per_hour, ambient_f)
            rows.append((
                "Still-air rise",
                f"+{air_rise:.1f} F/hr per {DEFAULT_ROOM_VOLUME_FT3:,.0f} ft3",
                "Ambient",
                "No-loss estimate",
            ))
            rows.append((
                "Cooling airflow",
                f"{cfm:.0f} CFM for +{REFERENCE_EXHAUST_RISE_F:.0f} F exhaust",
                "Ambient",
                "HVAC estimate",
            ))
        else:
            rows.append(("Ambient input", "Enter e.g. 72 or 22C", "", "Not used"))
        rows.append(("Session energy", format_kwh(self.session_wh / 1000.0), "Accumulated", "Derived"))
        rows.append(("Session heat", format_btu(self.session_wh * WATTS_TO_BTU_PER_HOUR), "Derived", "Derived"))
        for reading in snapshot.power:
            used = "Yes" if reading.selected_for_total else ""
            if reading.estimated:
                used = "Estimate"
            rows.append((reading.name, reading.display_value(), used, display_source(reading)))
        return rows

    def _temperature_rows(self, snapshot: HardwareSnapshot, ambient_f: Optional[float]) -> list[tuple[str, str, str]]:
        rows = [(reading.name, reading.display_value(), display_source(reading)) for reading in snapshot.temperatures]
        if snapshot.max_temp_c is not None:
            source = "Derived"
            if any(not reading.max_temperature_eligible for reading in snapshot.temperatures):
                source = "Derived (Pascal-adjusted)"
            rows.insert(0, ("Max temperature", format_temp(snapshot.max_temp_c), source))
            if ambient_f is not None:
                ambient_c = fahrenheit_to_celsius(ambient_f)
                rows.insert(1, ("Ambient air", format_temp_f_c(ambient_f), "User input"))
                rows.insert(2, ("Max above ambient", format_temp_delta(snapshot.max_temp_c - ambient_c), "Derived"))
            else:
                rows.insert(1, ("Ambient air", "Enter e.g. 72 or 22C", "Not used"))
        return rows

    def _status_note(self, snapshot: HardwareSnapshot, ambient_f: Optional[float]) -> str:
        if ambient_f is None:
            return "Enter ambient as Fahrenheit, or add C for Celsius, to unlock above-ambient and air-rise estimates."
        if snapshot.notes:
            return snapshot.notes[0]
        return "BTU/hr = watts x 3.412. Almost all PC electrical power becomes room heat."

    def on_close(self) -> None:
        self.stop_event.set()
        self.root.after(50, self.root.destroy)
def export_session_log_to_excel(
    session_log: list[SessionLogEntry],
    output_path: str,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Main data sheet
    ws = wb.active
    ws.title = "HeatLens Log"

    headers = [
        "Timestamp",
        "Elapsed Seconds",
        "Total Watts",
        "Direct Watts",
        "Estimated Watts",
        "Average kWh/hr",
        "BTU/hr",
        "Max Temp C",
        "Ambient F",
        "Heat Equivalent",
        "Status",
        "Selected Sources",
        "Notes",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for entry in session_log:
        ws.append([
            entry.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            round(entry.elapsed_seconds, 1),
            round(entry.total_watts, 2),
            round(entry.direct_watts, 2),
            round(entry.estimated_watts, 2),
            round(entry.average_kwh_per_hour, 4)
                if entry.average_kwh_per_hour is not None else None,
            round(entry.btu_per_hour, 2),
            round(entry.max_temp_c, 2)
                if entry.max_temp_c is not None else None,
            round(entry.ambient_f, 2)
                if entry.ambient_f is not None else None,
            entry.heat_equivalent,
            entry.status,
            entry.selected_sources,
            entry.notes,
        ])

    # Auto-size columns
    for column in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = min(max_len + 3, 60)

    # Summary sheet
    summary = wb.create_sheet("Summary")

    total_samples = len(session_log)
    avg_watts = (
        sum(e.total_watts for e in session_log) / total_samples
        if total_samples else 0
    )

    peak_watts = max((e.total_watts for e in session_log), default=0)
    peak_btu = max((e.btu_per_hour for e in session_log), default=0)

    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)

    summary.append(["Samples Logged", total_samples])
    summary.append(["Average Watts", round(avg_watts, 2)])
    summary.append(["Peak Watts", round(peak_watts, 2)])
    summary.append(["Peak BTU/hr", round(peak_btu, 2)])

    if session_log:
        duration_hours = (
            session_log[-1].elapsed_seconds / 3600.0
        )
        energy_kwh = (
            sum(e.total_watts for e in session_log)
            / total_samples
            * duration_hours
            / 1000.0
        )

        summary.append(["Duration Hours", round(duration_hours, 3)])
        summary.append(["Estimated Energy (kWh)", round(energy_kwh, 3)])

    wb.save(output_path)

def configure_ttk_style() -> None:
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass
    style.configure(
        "Treeview",
        background=COLORS["card"],
        foreground=COLORS["text"],
        fieldbackground=COLORS["card"],
        borderwidth=0,
        rowheight=26,
        font=("Segoe UI", 9),
    )
    style.configure(
        "Treeview.Heading",
        background=COLORS["card_2"],
        foreground=COLORS["muted"],
        relief="flat",
        font=("Segoe UI", 9, "bold"),
    )
    style.map("Treeview", background=[("selected", COLORS["card_2"])])
    style.configure(
        "HeatLens.TButton",
        background=COLORS["card_2"],
        foreground=COLORS["text"],
        bordercolor=COLORS["border"],
        focusthickness=0,
        padding=(12, 7),
        font=("Segoe UI", 9, "bold"),
    )
    style.map(
        "HeatLens.TButton",
        background=[("active", COLORS["border"]), ("pressed", COLORS["card"])],
        foreground=[("disabled", COLORS["subtle"])],
    )
    style.configure(
        "HeatLens.TEntry",
        fieldbackground=COLORS["card_2"],
        foreground=COLORS["text"],
        bordercolor=COLORS["border"],
        insertcolor=COLORS["text"],
        padding=(5, 4),
        font=("Segoe UI", 9),
    )
    style.configure(
        "HeatLens.TCheckbutton",
        background=COLORS["bg"],
        foreground=COLORS["text"],
        font=("Segoe UI", 9, "bold"),
    )
    style.map(
        "HeatLens.TCheckbutton",
        background=[("active", COLORS["bg"])],
        foreground=[("disabled", COLORS["subtle"])],
    )


def main() -> None:
    if maybe_elevate_on_startup():
        return
    set_dpi_awareness()
    root = tk.Tk()
    HeatLensApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
